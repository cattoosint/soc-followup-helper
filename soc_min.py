"""SOC follow-up - minimal edition (~150 lines, meant to be retyped by hand).

Reads case numbers from a sheet, searches each one in Outlook on the web,
opens the Reply All draft, and waits. It never sends anything - you review
and press Send in Outlook, then press Enter here for the next case.

    pip install selenium openpyxl
    python soc_min.py cases.xlsx
    python soc_min.py cases.xlsx --dry     (find only, opens no drafts)
"""

import csv
import re
import sys
import time
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

URL = "https://outlook.cloud.microsoft/mail/"
PROFILE = str(Path(__file__).resolve().parent / "profile")
NUM = re.compile(r"\d{5,8}")
ROWS = ("div[aria-label='Message list'] div[role='option']",
        "[aria-label*='Message list' i] [role='option']",
        "[role='listbox'] [role='option']")


def read_cases(path):
    """Case numbers from column 1 of a .xlsx (filtered rows skipped) or .csv."""
    out = []
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        ws = load_workbook(path, data_only=True).active
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            dim = ws.row_dimensions.get(i)
            if i == 1 or (dim is not None and dim.hidden):
                continue
            m = NUM.search(str(row[0]))
            if m and m.group() not in out:
                out.append(m.group())
    else:
        with open(path, encoding="utf-8-sig") as f:
            for row in list(csv.reader(f))[1:]:
                m = NUM.search(row[0]) if row else None
                if m and m.group() not in out:
                    out.append(m.group())
    return out


def find(driver, css_list):
    for css in css_list:
        for el in driver.find_elements(By.CSS_SELECTOR, css):
            if el.is_displayed():
                return el
    return None


def rows_for(driver, num):
    """Visible message-list rows whose text holds this exact case number."""
    pat = re.compile(rf"(?<!\d){num}(?!\d)")
    for css in ROWS:
        found = [e for e in driver.find_elements(By.CSS_SELECTOR, css)
                 if e.is_displayed()]
        hits = [e for e in found
                if pat.search(e.get_attribute("aria-label") or e.text or "")]
        if hits:
            return hits
    return []


def search(driver, query, num):
    box = find(driver, ["#topSearchInput", "input[aria-label*='Search' i]"])
    if box is None:
        return []
    box.click()
    box.send_keys(Keys.CONTROL, "a")
    box.send_keys(Keys.DELETE)
    box.send_keys(query)
    time.sleep(0.4)
    box.send_keys(Keys.ENTER)
    for _ in range(12):
        time.sleep(0.8)
        hits = rows_for(driver, num)
        if hits:
            return hits
    return []


def reply_all(driver):
    """Toolbar button, else the '...' overflow menu."""
    btn = find(driver, ["button[aria-label='Reply all']", "[aria-label='Reply all']"])
    if btn is None:
        more = find(driver, ["button[aria-label='More options']",
                             "button[aria-label='More actions']"])
        if more is None:
            return False
        more.click()
        time.sleep(0.8)
        for el in driver.find_elements(
                By.XPATH, "//*[@role='menuitem'][normalize-space(.)='Reply all']"):
            if el.is_displayed():
                el.click()
                return True
        return False
    btn.click()
    return True


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    dry = "--dry" in sys.argv
    if not args:
        sys.exit("usage: python soc_min.py cases.xlsx [--dry]")
    cases = read_cases(args[0])
    print(f"{len(cases)} case(s) to follow up")

    opts = Options()
    opts.add_argument(f"--user-data-dir={PROFILE}")
    opts.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=opts)
    driver.get(URL)
    input("Sign in to Outlook if asked, then press Enter here to start... ")

    results = []
    for i, num in enumerate(cases, 1):
        print(f"\n[{i}/{len(cases)}] SOC{num}")
        hits = []
        for q in (f"SOC{num}", f"SOC {num}", num):
            hits = search(driver, q, num)
            if hits:
                print(f"    found via {q!r}")
                break
        if not hits:
            print("    NOT FOUND - flagged")
            results.append((num, "NOT_FOUND"))
            continue
        if dry:
            results.append((num, "FOUND"))
            continue

        hits[0].click()
        time.sleep(2)
        if not reply_all(driver):
            print("    Reply all not found - do this one by hand")
            results.append((num, "ERROR"))
            continue
        print("    Draft open. Review and press Send in Outlook.")
        choice = input("    [Enter]=next case | s=skip | q=quit > ").strip().lower()
        if choice == "q":
            results.append((num, "QUIT"))
            break
        results.append((num, "SKIPPED" if choice == "s" else "SENT"))
        time.sleep(5)

    out = Path(f"followup_results_{time.strftime('%Y%m%d_%H%M%S')}.csv")
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows([("case", "status")] +
                                [(f"SOC{n}", s) for n, s in results])
    missing = [n for n, s in results if s in ("NOT_FOUND", "ERROR")]
    print(f"\nDone. {len(results)} processed, {len(missing)} need manual "
          f"follow-up. Log: {out.name}")
    for n in missing:
        print(f"   SOC{n}")
    driver.quit()


if __name__ == "__main__":
    main()
