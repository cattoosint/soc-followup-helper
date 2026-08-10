#!/usr/bin/env python3
"""SOC nightshift follow-up helper (core + command line).

Drives Outlook on the web only - the SOC uses OWA, so there is no Outlook
desktop/COM path to maintain.

Reads case numbers from an XSOAR export (.csv or .xlsx), then drives Outlook
on the web: for each case it searches  subject:SOC<num>  (and
subject:"SOC <num>"  as a fallback), opens the newest matching mail, clicks
Reply all, and waits for the analyst to review + send before moving on.
Cases with no matching mail are flagged and written to a results CSV plus a
highlighted copy of the input sheet.

.xlsx input: rows hidden by an Excel filter are SKIPPED, so "filter, save,
upload" processes only the visible rows. (CSV can't record filters - if you
filter in Excel, save as .xlsx.)

CLI usage:
    python soc_followup.py --csv cases.xlsx
    python soc_followup.py --csv cases.csv --parse-only
GUI:
    python soc_gui.py
"""

import argparse
import csv
import logging
import os
import platform
import re
import sys
import time
import traceback

# Use the libraries bundled in libs\ if they are present, so the tool runs
# straight from the folder with nothing to pip install. A system-installed
# selenium/openpyxl still wins if the analyst has one.
_LIBS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "libs")
if os.path.isdir(_LIBS) and _LIBS not in sys.path:
    sys.path.append(_LIBS)
from datetime import datetime, timedelta
from pathlib import Path

try:  # selenium ships with seleniumbase; absent until deps are installed
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
except ImportError:
    By = Keys = None

SCRIPT_DIR = Path(__file__).resolve().parent
CASE_NUM_RE = re.compile(r"\d{5,8}")
COLUMN_HINTS = ("case", "incident", "ticket", "id", "number")
NO_RESULTS_RE = re.compile(
    r"(couldn.t find|didn.t find|no results|nothing (?:that )?match)", re.I
)

# ------------------------------------------------------------------- settings

SETTINGS_PATH = SCRIPT_DIR / "soc_settings.json"


def load_settings():
    """Local preferences (mailbox, auto-send sender, ...).

    Kept in soc_settings.json beside the tool and never committed - the
    sender address is site-specific and does not belong in a repo.
    """
    import json
    try:
        return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_settings(values):
    import json
    try:
        SETTINGS_PATH.write_text(json.dumps(values, indent=2), encoding="utf-8")
        return True
    except Exception as e:
        log.debug("could not save settings: %s", e)
        return False


# -------------------------------------------------------------------- logging

log = logging.getLogger("soc_followup")
LOG_PATH = None


def setup_logging(verbose=False):
    """Write a detailed log to logs\\followup_<timestamp>.log.

    The analyst can send this file on when something goes wrong - it holds
    the environment, every search tried, what matched, and full tracebacks.
    """
    global LOG_PATH
    if LOG_PATH:
        return LOG_PATH
    logs = SCRIPT_DIR / "logs"
    try:
        logs.mkdir(exist_ok=True)
    except Exception:
        return None
    LOG_PATH = logs / f"followup_{datetime.now():%Y%m%d_%H%M%S}.log"
    handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    handler.setFormatter(logging.Formatter(
        "%(asctime)s %(levelname)-7s %(message)s", "%H:%M:%S"))
    log.setLevel(logging.DEBUG)
    log.addHandler(handler)
    if verbose:
        stream = logging.StreamHandler()
        stream.setFormatter(logging.Formatter("    [log] %(message)s"))
        log.addHandler(stream)
    log.info("=" * 60)
    log.info("SOC nightshift follow-up - new run")
    log.info("=" * 60)
    for line in environment_report():
        log.info(line)
    return LOG_PATH


def environment_report():
    """Everything needed to reproduce a problem remotely."""
    out = [f"tool folder   : {SCRIPT_DIR}",
           f"python        : {sys.version.split()[0]} ({sys.executable})",
           f"platform      : {platform.platform()}"]
    out.append(f"bundled libs  : {'yes' if os.path.isdir(_LIBS) else 'no'}")
    for mod in ("selenium", "openpyxl", "psutil", "seleniumbase"):
        try:
            m = __import__(mod)
            where = "bundled" if _LIBS in str(getattr(m, "__file__", "")) else "system"
            out.append(f"{mod:<14}: {getattr(m, '__version__', 'installed')} ({where})")
        except ImportError:
            out.append(f"{mod:<14}: not installed")
    return out


def log_exception(context, exc):
    log.error("%s: %s: %s", context, type(exc).__name__, exc)
    log.debug("traceback:\n%s", traceback.format_exc())


def log_browser(driver):
    """Chrome/driver versions and the window state, once at startup."""
    try:
        caps = getattr(driver, "capabilities", {}) or {}
        log.info("chrome        : %s", caps.get("browserVersion", "?"))
        chrome = caps.get("chrome", {}) or {}
        log.info("chromedriver  : %s",
                 str(chrome.get("chromedriverVersion", "?")).split(" ")[0])
        log.info("window        : %s", driver.get_window_rect())
        log.info("start url     : %s", driver.current_url)
    except Exception as e:
        log.debug("couldn't read browser info: %s", e)


# ---------------------------------------------------------------- file input

def _unique_headers(headers):
    """Make repeated column names distinct.

    Exports sometimes carry the same heading twice. Zipping them into a
    dict would silently keep only the last one, so the tool could read
    case numbers from the wrong column.
    """
    seen, out = {}, []
    for i, h in enumerate(headers):
        name = str(h).strip() if h is not None and str(h).strip() else f"column{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        out.append(name)
    return out


def read_csv_rows(path, stats=None):
    """Read a .csv or .xlsx file into (headers, list-of-dict-rows).

    For .xlsx, rows hidden by a filter are skipped (count reported via
    stats["hidden_skipped"] if a dict is passed).
    """
    path = Path(str(path).strip().strip('"').strip("'"))
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("reading .xlsx needs openpyxl - run: pip install openpyxl")
        try:
            book = load_workbook(path, data_only=True)
            ws = book.active
            if len(book.sheetnames) > 1:
                # .active is whichever tab was showing when it was saved
                log.info("workbook has %d sheets %s - reading the active one, "
                         "%r", len(book.sheetnames), book.sheetnames, ws.title)
                if stats is not None:
                    stats["sheets"] = book.sheetnames
                    stats["sheet_used"] = ws.title
        except PermissionError:
            # open in Excel: work from a copy rather than making the analyst
            # close their sheet
            import shutil
            import tempfile
            tmp = Path(tempfile.gettempdir()) / f"_soc_followup_{path.name}"
            shutil.copy2(path, tmp)
            ws = load_workbook(tmp, data_only=True).active
        hidden = 0
        data, line_numbers = [], []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > 1 and idx in ws.row_dimensions and ws.row_dimensions[idx].hidden:
                hidden += 1
                continue
            data.append([("" if c is None else str(c)) for c in row])
            line_numbers.append(idx)   # the row number Excel shows
        if stats is not None:
            stats["hidden_skipped"] = hidden
            stats["line_numbers"] = line_numbers[1:]   # drop the heading row
        if not data:
            return [], []
        headers = _unique_headers(data[0])
        rows = [dict(zip(headers, r)) for r in data[1:]]
        return headers, rows
    source = path
    try:
        open(path, "rb").close()
    except PermissionError:
        import shutil
        import tempfile
        source = Path(tempfile.gettempdir()) / f"_soc_followup_{path.name}"
        shutil.copy2(path, source)

    # Excel writes CSV in the local codepage, not UTF-8, so a name with an
    # accent in it would otherwise make the whole export unreadable
    rows, headers = [], []
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(source, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                rows = [row for row in reader]
                headers = reader.fieldnames or []
            if encoding != "utf-8-sig":
                log.info("read %s as %s", path.name, encoding)
            break
        except UnicodeDecodeError:
            continue
    headers = _unique_headers(headers)
    rows = [{h: r.get(o) for h, o in zip(headers, r.keys())} for r in rows] \
        if rows and len(headers) == len(rows[0]) else rows
    if stats is not None:
        stats["hidden_skipped"] = 0
        stats["line_numbers"] = list(range(2, len(rows) + 2))
    return headers, rows


def detect_case_column(headers, rows):
    """Pick the column most likely to hold case numbers."""
    best_col, best_score = None, 0.0
    total = len(rows) or 1
    for col in headers:
        if not col:
            continue
        values = [str(r.get(col, "") or "").strip() for r in rows]
        values = [v for v in values if v]
        if not values:
            continue
        hits = sum(1 for v in values if CASE_NUM_RE.search(v))
        # score against EVERY row, not just the filled ones - otherwise a
        # column with two stray numbers and 300 blanks scores a perfect 1.0
        # and beats the real id column
        frac = hits / total
        named_like_a_case = any(h in col.lower() for h in COLUMN_HINTS)
        bonus = 0.5 if named_like_a_case else 0.0
        # A column called "Case ID" counts even if some rows are unparseable -
        # a messy export shouldn't stop the tool dead. The rows it can't read
        # are reported by preflight() rather than silently dropped.
        threshold = 0.3 if named_like_a_case else 0.6
        if frac >= threshold and frac + bonus > best_score:
            best_col, best_score = col, frac + bonus
    return best_col


def extract_cases(path, column=None, stats=None):
    """Returns (column_used, [case numbers as strings, deduped, in order])."""
    headers, rows = read_csv_rows(path, stats)
    if not headers:
        raise ValueError(f"no header row found in {path}")
    if column:
        if column not in headers:
            raise ValueError(
                f"column '{column}' not in file. Available: {', '.join(headers)}"
            )
    else:
        column = detect_case_column(headers, rows)
        if not column:
            raise ValueError(
                "could not auto-detect the case-number column. "
                f"Available columns: {', '.join(headers)}"
            )
    cases, seen = [], set()
    unreadable, duplicates, blanks = [], [], []
    # report the row numbers Excel shows, which skip over filtered-out rows
    line_numbers = (stats or {}).get("line_numbers") or \
        list(range(2, len(rows) + 2))
    for i, row in zip(line_numbers, rows):
        raw = str(row.get(column) or "").strip()
        m = CASE_NUM_RE.search(raw)
        if not m:
            if raw:
                unreadable.append((i, raw[:40]))
            else:
                blanks.append(i)                # empty cell - still a dropped row
            continue
        if m.group(0) in seen:
            duplicates.append((i, m.group(0)))
            continue
        seen.add(m.group(0))
        cases.append(m.group(0))
    if stats is not None:
        stats["rows_read"] = len(rows)
        stats["unreadable"] = unreadable        # rows that hold no case number
        stats["duplicates"] = duplicates
        stats["blanks"] = blanks
        stats["duplicate_lines"] = {i for i, _ in duplicates}
    return column, cases


def preflight(path, column=None):
    """Dry-run an export: what will be processed, and what will be ignored.
    Nothing touches Outlook - use before a shift to sanity-check a new export.
    """
    stats = {}
    col, cases = extract_cases(path, column, stats)
    headers, _ = read_csv_rows(path)
    # An export with no header row would have its first case number read as a
    # column name - i.e. one case silently missing from the whole run.
    header_is_data = [str(h) for h in headers
                      if h and CASE_NUM_RE.search(str(h))]
    return {
        "column": col,
        "columns": headers,
        "cases": cases,
        "rows_read": stats.get("rows_read", 0),
        "hidden_skipped": stats.get("hidden_skipped", 0),
        "unreadable": stats.get("unreadable", []),
        "duplicates": stats.get("duplicates", []),
        "blanks": stats.get("blanks", []),
        "duplicate_lines": stats.get("duplicate_lines", set()),
        "header_is_data": header_is_data,
        "sheets": stats.get("sheets", []),
        "sheet_used": stats.get("sheet_used", ""),
    }


def format_preflight(report):
    out = [f"Case-number column : '{report['column']}'",
           f"Columns in file    : {', '.join(str(c) for c in report['columns'])}",
           f"Rows read          : {report['rows_read']}"]
    if len(report.get("sheets") or []) > 1:
        out.append(f"!! The workbook has {len(report['sheets'])} sheets "
                   f"{report['sheets']}.")
        out.append(f"   Reading {report['sheet_used']!r} - the one that was "
                   "showing when it was saved.")
    if report["hidden_skipped"]:
        out.append(f"Filtered out (Excel): {report['hidden_skipped']} row(s) skipped")
    out.append(f"Cases to process   : {len(report['cases'])}")
    if report["duplicates"]:
        out.append(f"Duplicates ignored : {len(report['duplicates'])} "
                   f"(e.g. row {report['duplicates'][0][0]}: "
                   f"{report['duplicates'][0][1]})")
    if report.get("blanks"):
        out.append(f"Blank case cells   : {len(report['blanks'])} row(s) "
                   f"(e.g. row {report['blanks'][0]}) - ignored")
    if report["unreadable"]:
        out.append(f"!! NO CASE NUMBER FOUND in {len(report['unreadable'])} row(s) "
                   "- these will be IGNORED:")
        for line, raw in report["unreadable"][:10]:
            out.append(f"     row {line}: {raw!r}")
        if len(report["unreadable"]) > 10:
            out.append(f"     ... and {len(report['unreadable']) - 10} more")
        out.append("   (case numbers must be 5-8 digits; tell me the real "
                   "format if yours differ)")
    if report.get("header_is_data"):
        out.append("!! The first row looks like DATA, not column headings "
                   f"({', '.join(report['header_is_data'][:3])}).")
        out.append("   Row 1 is always treated as headings, so that case would "
                   "be missed - add a heading row to the export.")
    sample = ", ".join(f"SOC{c}" for c in report["cases"][:6])
    if sample:
        out.append(f"First few          : {sample}")
    return "\n".join(out)

# --------------------------------------------------------------- UI plumbing

class ConsoleUI:
    """Terminal front-end: log() prints, ask() uses input()."""

    def log(self, msg):
        print(msg)

    def ask(self, prompt, choices, kind=None):
        while True:
            ans = input(prompt).strip().lower()
            if ans in choices:
                return ans
            print("    Unrecognised option.")

    def case_update(self, num, status, detail=""):
        pass  # GUIs use this to move cases through their tracker table

    def ask_or_watch(self, prompt, choices, watch_fn, poll_s=0.7, kind=None):
        """Like ask(), but also auto-returns 'auto' once watch_fn() is True
        (e.g. the reply draft was sent). Falls back to blocking ask() if
        single-key console input isn't available."""
        print(prompt + " (auto-continues once the draft is sent)")
        try:
            import msvcrt
        except ImportError:
            return self.ask(prompt, choices)
        while True:
            if watch_fn():
                print("    Send detected - continuing.")
                return "auto"
            t = time.time() + poll_s
            while time.time() < t:
                if msvcrt.kbhit():
                    ch = msvcrt.getwch()
                    if ch in ("\r", "\n"):
                        return ""
                    if ch.lower() in choices:
                        return ch.lower()
                time.sleep(0.05)


def _ask_with_watch(ui, prompt, choices, watch_fn, kind=None):
    fn = getattr(ui, "ask_or_watch", None)
    if fn and watch_fn:
        return fn(prompt, choices, watch_fn, kind=kind)
    return ui.ask(prompt, choices, kind=kind)


def _notify(ui, num, status, detail=""):
    log.info("case %s -> %s %s", num, status, f"({detail})" if detail else "")
    fn = getattr(ui, "case_update", None)
    if fn:
        try:
            fn(num, status, detail)
        except Exception:
            pass

# ------------------------------------------------------------- Outlook (OWA)

def _find_first(driver, css_list):
    """First displayed element matching any of the CSS selectors, else None."""
    for css in css_list:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, css)
        except Exception:
            continue
        for el in els:
            try:
                if el.is_displayed():
                    return el
            except Exception:
                continue
    return None


def get_search_box(driver):
    return _find_first(driver, [
        "#topSearchInput",
        "input[role='searchbox']",
        "[role='searchbox']",
        "input[aria-label*='Search' i]",
        "input[placeholder*='Search' i]",
    ])


LOGIN_URL_HINTS = ("login.microsoftonline.com", "login.live.com",
                   "login.microsoft", "/oauth", "signin")


def on_login_page(page):
    """Are we sitting on a Microsoft sign-in page rather than the mailbox?"""
    try:
        url = (page.current_url or "").lower()
        if any(hint in url for hint in LOGIN_URL_HINTS):
            return True
        return _find_first(page, ["input[type='password']",
                                  "input[name='loginfmt']", "#i0116"]) is not None
    except Exception:
        return False


def wait_for_mailbox(page, timeout_s, ui=None):
    """Wait until the inbox is actually open.

    Signing in can take a while - MFA, a password prompt, picking an
    account - so this keeps waiting and says what it is waiting for rather
    than failing while the analyst is still typing.
    """
    deadline = time.time() + timeout_s
    announced = False
    next_nudge = time.time() + 30
    while time.time() < deadline:
        box = get_search_box(page)
        if box is not None:
            if announced:
                log.info("signed in - inbox reached")
                if ui:
                    ui.log("    Signed in - inbox loaded, carrying on.")
            return box
        if not announced and on_login_page(page):
            announced = True
            log.info("sign-in page detected - waiting for the analyst")
            if ui:
                ui.log("    Sign-in page - log in in the browser window "
                       "(tick 'Stay signed in?'). Waiting for the inbox...")
        if time.time() >= next_nudge:
            next_nudge = time.time() + 30
            left = int(deadline - time.time())
            log.debug("still waiting for the inbox (%ss left)", left)
            if ui and announced:
                ui.log(f"    Still waiting for sign-in... ({left // 60} min left)")
        time.sleep(2)
    log.warning("gave up waiting for the inbox after %ss", timeout_s)
    return None


def build_queries(num):
    """Search terms to try, in order, for one case number.

    Spacing matters, case does not - both proven live on 2026-08-04:
    "SOC610454" found nothing while "SOC 610454" matched, and "SOC610529"
    (uppercase) matched a mail whose subject is lowercase "soc610529". So
    both spacings are tried, but lowercase duplicates are not - they would
    only slow down cases that have no mail at all. The subject: operator is
    last; it matched nothing live.
    """
    return [
        f"SOC{num}",            # SOC610529  - the usual SOC format
        f"SOC {num}",           # SOC 610529 - and every punctuated form:
                                # proven live that this also finds
                                # [SOC#610529] and SOC-610529, because
                                # Outlook splits words on punctuation
        str(num),               # 610529     - bare number
        f"subject:SOC{num}",
        f'subject:"SOC {num}"',
    ]


_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*([AP])\.?M\.?", re.I)
_TIME24_RE = re.compile(r"\b([01]?\d|2[0-3]):([0-5]\d)\b")
_FULLDATE_RE = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b")
_SHORTDATE_RE = re.compile(r"\b(\d{1,2})/(\d{1,2})\b")
# only real weekday words - "\w*" used to swallow "Monitoring", "Satellite",
# "Wedge" and friends, back-dating a mail by up to six days
_WEEKDAY_RE = re.compile(
    r"\b(Mon|Monday|Tue|Tues|Tuesday|Wed|Wednesday|Thu|Thur|Thurs|Thursday|"
    r"Fri|Friday|Sat|Saturday|Sun|Sunday)\b", re.I)
_TODAY_RE = re.compile(r"\btoday\b", re.I)
_YESTERDAY_RE = re.compile(r"\byesterday\b", re.I)
_WEEKDAYS = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}


_DAY_FIRST = None


def date_is_day_first():
    """Does this machine write dates as d/m/yyyy rather than m/d/yyyy?

    Outlook follows the Windows short-date setting, so "10/8/2026" means
    10 August here and 8 October in the US. Guessing wrong reorders mail by
    months, which decides which one gets replied to.
    """
    global _DAY_FIRST
    if _DAY_FIRST is None:
        _DAY_FIRST = False
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                r"Control Panel\International") as k:
                fmt = winreg.QueryValueEx(k, "sShortDate")[0]
            _DAY_FIRST = fmt.strip().lower().startswith("d")
            log.debug("windows short date %r -> day first: %s", fmt, _DAY_FIRST)
        except Exception:
            pass
    return _DAY_FIRST


def _day_month(first, second):
    """(day, month) from two ambiguous date parts."""
    if first > 12:
        return first, second        # only a day can exceed 12
    if second > 12:
        return second, first
    return (first, second) if date_is_day_first() else (second, first)


def parse_row_time(label, now=None):
    """Best-effort timestamp from an Outlook row label.

    Handles the formats OWA actually shows: '7:34 PM' (today),
    'Mon 3:12 PM' (this week), 'Sat 7/11' (this year), '7/4/2026'.
    Returns None when nothing parseable is present.
    """
    if not label:
        return None
    now = now or datetime.now()
    hour = minute = None
    t = _TIME_RE.search(label)
    if t:
        hour = int(t.group(1)) % 12
        if t.group(3).upper() == "P":
            hour += 12
        minute = int(t.group(2))
    else:
        t24 = _TIME24_RE.search(label)      # mailboxes set to a 24-hour clock
        if t24:
            hour, minute = int(t24.group(1)), int(t24.group(2))

    if _YESTERDAY_RE.search(label):
        day = now - timedelta(days=1)
        return day.replace(hour=hour or 0, minute=minute or 0,
                           second=0, microsecond=0)
    if _TODAY_RE.search(label):
        return now.replace(hour=hour or 0, minute=minute or 0,
                           second=0, microsecond=0)

    m = _FULLDATE_RE.search(label)
    if m:
        d, mo = _day_month(int(m.group(1)), int(m.group(2)))
        y = int(m.group(3))
        if y < 100:
            y += 2000
        try:
            base = datetime(y, mo, d)
        except ValueError:
            return None
        return base.replace(hour=hour or 0, minute=minute or 0)

    m = _SHORTDATE_RE.search(label)
    if m:
        d, mo = _day_month(int(m.group(1)), int(m.group(2)))
        for year in (now.year, now.year - 1):
            try:
                base = datetime(year, mo, d, hour or 0, minute or 0)
            except ValueError:
                return None
            if base <= now + timedelta(days=1):
                return base
        return None

    if hour is None:
        return None

    wd = _WEEKDAY_RE.search(label)
    day = now
    if wd:
        target = _WEEKDAYS[wd.group(1)[:3].lower()]
        delta = (now.weekday() - target) % 7
        # Outlook labels a mail by weekday only when it is not today, so a
        # matching weekday name means a week ago, not this morning
        day = now - timedelta(days=delta or 7)
    stamp = day.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if stamp > now + timedelta(minutes=5):
        # a time later today means it arrived yesterday, not in the future
        stamp -= timedelta(days=1)
    return stamp


def sort_newest_first(items, now=None):
    """Order result rows newest-first by their displayed timestamp.
    Rows with no parseable time keep their original order, after the rest."""
    now = now or datetime.now()
    decorated = []
    for i, el in enumerate(items):
        ts = parse_row_time(label_of(el), now)
        decorated.append((0 if ts else 1, -ts.timestamp() if ts else i, i, el))
    decorated.sort(key=lambda d: d[:3])
    return [d[3] for d in decorated]


# The search box's autocomplete dropdown uses the same listbox/option roles
# as the message list, so a bare "[role='option']" also matches suggestions -
# clicking one runs a search instead of opening mail. Prefer the message list.
MESSAGE_LIST_SELECTORS = [
    "div[aria-label='Message list'] div[role='option']",
    "[aria-label*='Message list' i] [role='option']",
    "[data-app-section='MessageList'] [role='option']",
    "[role='listbox'][aria-label*='essage'] [role='option']",
    "div[role='listbox'] div[role='option'][data-convid]",
    "[role='listbox'] [role='option']",  # last resort
]


def _visible_results(page):
    last = len(MESSAGE_LIST_SELECTORS) - 1
    for i, css in enumerate(MESSAGE_LIST_SELECTORS):
        try:
            found = [e for e in page.find_elements(By.CSS_SELECTOR, css)
                     if e.is_displayed()]
        except Exception:
            continue
        if i == last:
            # Generic selector: also matches autocomplete suggestions. Every
            # real mail row shows a date/time, suggestions don't - so require
            # one rather than risk clicking a suggestion.
            found = [e for e in found if parse_row_time(label_of(e)) is not None]
        if found:
            return found
    return []


def _dismiss_suggestions(page):
    """Close the search autocomplete popup so it can't be mistaken for
    results (and can't swallow the next click)."""
    try:
        popup = page.find_elements(
            By.CSS_SELECTOR, "[role='listbox'][aria-label*='uggest' i], "
                             "[role='listbox'] [role='option'][id*='uggest']")
        if any(e.is_displayed() for e in popup):
            page.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
            time.sleep(0.4)
    except Exception:
        pass


def label_matches_case(label, num):
    """True only if `num` appears as a whole number in `label`.

    A plain substring test would make case 610529 match a mail about
    SOC1610529 - a different case entirely - so digits either side
    disqualify the match.
    """
    if not label:
        return False
    return re.search(rf"(?<!\d){re.escape(str(num))}(?!\d)", label) is not None


def label_of(el):
    try:
        return " ".join(((el.get_attribute("aria-label") or el.text) or "").split())
    except Exception:
        return ""


def run_search(page, query, settle_s, num=None):
    """Search OWA and return the visible result elements, or None.

    When num is given, only results whose row text actually contains that
    case number count as a match - so a search that silently returned the
    whole inbox doesn't get mistaken for a hit.
    """
    box = get_search_box(page)
    if box is None:
        raise RuntimeError("search box not found - are you logged in?")
    try:
        box.click()
    except Exception:
        pass
    box = get_search_box(page) or box  # OWA can swap the input once focused
    try:
        box.send_keys(Keys.CONTROL, "a")
        box.send_keys(Keys.DELETE)
    except Exception:
        pass
    for ch in query:  # human-ish typing pace
        box.send_keys(ch)
        time.sleep(0.03)
    time.sleep(0.4)       # let autocomplete settle so Enter submits what we typed
    box.send_keys(Keys.ENTER)
    time.sleep(settle_s)  # let the result list replace the inbox list
    _dismiss_suggestions(page)

    deadline = time.time() + 12
    misses = 0
    while time.time() < deadline:
        items = _visible_results(page)
        if items:
            if num is None:
                return items
            hits = [e for e in items if label_matches_case(label_of(e), num)]
            if hits:
                # let the list settle: results stream in, and returning on
                # the first poll can miss a newer mail arriving a moment
                # later - which is the one that should be replied to
                time.sleep(0.8)
                settled = [e for e in _visible_results(page)
                           if label_matches_case(label_of(e), num)]
                if len(settled) > len(hits):
                    log.debug("result list grew from %d to %d while settling",
                              len(hits), len(settled))
                    hits = settled
                log.info("search %r -> %d row(s), %d match case %s",
                         query, len(items), len(hits), num)
                log.debug("  matched rows: %s",
                          [label_of(e)[:90] for e in hits[:3]])
                return hits
            misses += 1
            if misses >= 3:  # results are showing, none are this case
                log.info("search %r -> %d row(s), none contain %s",
                         query, len(items), num)
                log.debug("  rows shown: %s",
                          [label_of(e)[:90] for e in items[:3]])
                return None
        else:
            try:
                body = page.find_element(By.TAG_NAME, "body").text
                if NO_RESULTS_RE.search(body or ""):
                    log.info("search %r -> Outlook reports no results", query)
                    return None
            except Exception:
                pass
        time.sleep(0.6)
    log.info("search %r -> nothing appeared within the timeout", query)
    return None


def _click_el(el):
    if el is None:
        return False
    try:
        if not el.is_enabled():
            return False
        el.click()
        return True
    except Exception:
        return False


def _find_by_text(page, text, roles=("menuitem", "button", "menuitemcheckbox")):
    """Visible element with one of these ARIA roles whose text is exactly
    `text` (used for menu entries that carry no aria-label)."""
    parts = [f"//*[@role='{r}'][normalize-space(.)='{text}']" for r in roles]
    parts.append(f"//button[normalize-space(.)='{text}']")
    try:
        for el in page.find_elements(By.XPATH, " | ".join(parts)):
            if el.is_displayed():
                return el
    except Exception:
        pass
    return None


def click_reply_all(page):
    """Click Reply all - toolbar button first, then the '...' overflow menu.

    On narrow windows OWA hides Reply all behind the '...' (More actions)
    button next to the message header, so the direct button isn't there.
    """
    direct = _find_first(page, [
        "button[aria-label='Reply all']",
        "[aria-label='Reply all']",
        "button[title='Reply all']",
        "[aria-label*='Reply all']",
    ])
    if _click_el(direct):
        return True

    more = _find_first(page, [
        "button[aria-label='More options']",
        "button[aria-label='More actions']",
        "button[aria-label='More commands']",
        "[aria-label='More options']",
        "[aria-label*='More actions' i]",
        "[data-testid='ThreeDotButton']",
    ])
    if not _click_el(more):
        return False
    time.sleep(0.8)

    if _click_el(_find_by_text(page, "Reply all")):
        return True
    if _click_el(_find_first(page, ["[aria-label='Reply all']"])):
        return True

    # some builds tuck it under "Other reply actions"
    other = _find_by_text(page, "Other reply actions")
    if _click_el(other):
        time.sleep(0.8)
        if _click_el(_find_by_text(page, "Reply all")):
            return True
    try:  # leave no menu hanging open for the next case
        page.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
    except Exception:
        pass
    return False


# A conversation stacks messages oldest at the top, newest at the bottom,
# so the newest message header is the lowest one on screen. Drafts sit at
# the very bottom and must not be mistaken for a sent message.
DRAFT_MARKERS = ("[draft]", "this message hasn't been sent", "saved:")
                         # 12-hour, 24-hour, or a date - a mailbox set to a
                         # 24-hour clock used to have no recognisable message
                         # headers at all, so auto-send could never fire
MSG_TIME_RE = re.compile(r"\d{1,2}:\d{2}\s*[AP]\.?M\.?|\b\d{1,2}:\d{2}\b|"
                         r"\d{1,2}/\d{1,2}(?:/\d{2,4})?", re.I)


def normalize_text(value):
    """Lowercase, with spacing and icon glyphs removed.

    Outlook breaks long addresses for wrapping, so a header can read
    "soc alertstest@outlook.com" - matching has to ignore that, and the
    icon characters the UI mixes into the same text. Written with escapes
    rather than literal characters: this file has been mangled before by a
    tool that rewrote it in a non-UTF8 encoding.
    """
    out = []
    for ch in (value or ""):
        if ch.isspace():
            continue
        if "\ue000" <= ch <= "\uf8ff":      # private-use icon glyphs
            continue
        if ch in ("\u200b", "\ufeff", "\u00a0"):   # zero-width / nbsp
            continue
        out.append(ch)
    return "".join(out).lower()


QUOTED_HEADER_RE = re.compile(r"\b(from|sent|subject|to|cc|bcc)\s*:", re.I)


def sender_part(header):
    """Just the sender of a message, from its header text.

    OWA renders a header as "<sender> To:<recipients> <date> ..." and the
    element text can also carry body content, including quoted headers from
    an earlier mail ("From: SOC Alerts <alerts@...> Sent: ..."). Matching
    against any of that would let a forwarded or quoted message pass as if
    the expected sender had written it, so anything at or after the first
    recipient/quoted-header marker is dropped, and so is anything after the
    first timestamp - the sender always precedes both.

    Returns "" when no plausible sender line can be isolated; callers must
    treat that as "do not auto-send".
    """
    text = " ".join((header or "").split())
    if not text:
        return ""

    cut = QUOTED_HEADER_RE.search(text)
    if cut:
        if cut.start() == 0:        # starts with "From:" - a quoted header,
            return ""               # not a real OWA sender line
        text = text[:cut.start()]

    stamp = MSG_TIME_RE.search(text)
    if stamp:
        text = text[:stamp.start()]

    return text.strip()[:120]


def reading_pane(page):
    for css in ("[aria-label*='Reading pane' i]", "#ReadingPaneContainerId",
                "div[role='document']"):
        try:
            el = page.find_element(By.CSS_SELECTOR, css)
            if el.is_displayed():
                return el
        except Exception:
            continue
    return None


def read_last_message(page):
    """(header, body) for the NEWEST message in the open conversation.

    header is that message's sender line - the one auto-send is judged on,
    so a thread whose latest reply came from someone else is never treated
    as if it came from the expected sender. Returns ("", "") when the page
    cannot be read, which callers must treat as "do not auto-send".
    """
    root = reading_pane(page)
    if root is None:
        return "", ""
    try:
        body = root.text or ""
    except Exception:
        return "", ""

    candidates = []
    for css in ("div[aria-label*='message' i]", "div[role='listitem']"):
        try:
            elements = root.find_elements(By.CSS_SELECTOR, css)
        except Exception:
            continue
        for el in elements:
            try:
                if not el.is_displayed():
                    continue
                text = " ".join((el.text or "").split())
                if not text or not MSG_TIME_RE.search(text):
                    continue          # not a message header
                if any(m in text.lower() for m in DRAFT_MARKERS):
                    continue          # an unsent draft, not a real message
                y = el.location.get("y", 0)
            except Exception:
                continue
            candidates.append((y, text, parse_row_time(text)))

    if not candidates:
        return "", body

    # Prefer the newest by its own timestamp. Screen position is not
    # reliable: "Show newest messages on top" is a per-mailbox setting, and
    # under it the bottom-most message is the OLDEST one.
    timed = [c for c in candidates if c[2] is not None]
    if timed:
        newest = max(timed, key=lambda c: c[2])
        ties = [c for c in timed if c[2] == newest[2]]
        if len(ties) > 1:                    # same minute - fall back to order
            newest = max(ties, key=lambda c: c[0])
        header = newest[1]
    elif len(candidates) == 1:
        header = candidates[0][1]
    else:
        # several messages, none with a readable time: cannot tell which is
        # newest, so say so rather than guess
        log.warning("could not order %d messages in the thread by time",
                    len(candidates))
        return "", body

    return sender_part(header), body


def should_auto_send(last_header, body, keyword, expected_sender):
    """Both conditions must hold, or the analyst handles it by hand.

    last_header is the newest message's sender line (see read_last_message).
    The sender may be given as an address or a display name - whichever the
    mailbox shows. Returns (send_it, reason_for_the_log).
    """
    if not keyword or not expected_sender:
        return False, "auto-send needs both a phrase and a sender configured"
    if not body:
        return False, "could not read the message"
    if not last_header:
        return False, "could not identify who sent the latest message"

    # compare with spacing removed: Outlook wraps long addresses mid-word
    # and sprinkles icon characters through the same text
    keyword_ok = normalize_text(keyword) in normalize_text(body)
    if len(normalize_text(expected_sender)) < 4:
        return False, ("the configured sender is too short to match safely "
                       "- use the full address or display name")
    # judged on the latest message's sender line only - never the whole
    # quoted thread, where the expected sender may appear from earlier mails
    sender_ok = normalize_text(expected_sender) in normalize_text(last_header)

    if keyword_ok and sender_ok:
        return True, f"matched {keyword!r}; latest mail is from {expected_sender!r}"
    missing = []
    if not keyword_ok:
        missing.append(f"{keyword!r} not in the message")
    if not sender_ok:
        missing.append(f"latest mail is not from {expected_sender!r} "
                       f"(saw {last_header[:70]!r})")
    return False, "; ".join(missing)


def compose_subject(page):
    """Subject line of the open draft, if it can be read."""
    el = _find_first(page, ["input[aria-label='Subject']",
                            "input[placeholder='Add a subject']",
                            "input[aria-label*='subject' i]"])
    if el is None:
        return ""
    try:
        return (el.get_attribute("value") or el.get_attribute("title")
                or el.text or "")
    except Exception:
        return ""


def click_send(page, num=None):
    """Press Send on the open draft.

    With `num`, the draft's subject must carry that case number first. The
    Send button is found by DOM order, so without this an unrelated draft -
    one left open from an earlier case - could be the one that goes out.
    """
    if num is not None:
        subject = compose_subject(page)
        if subject and not label_matches_case(subject, num):
            log.error("refusing to send: draft subject %r is not case %s",
                      subject[:80], num)
            return False
        if not subject:
            log.warning("could not read the draft subject for %s", num)
    el = _find_first(page, ["button[aria-label='Send']",
                            "button[aria-label^='Send (']",
                            "button[title^='Send']"])
    return _click_el(el)


def wait_for_reading_pane(page, timeout=12, num=None):
    """True once the clicked mail is open.

    When `num` is given, the pane must also be showing THAT case - the
    previous case's mail is still on screen for a moment after clicking,
    and acting on it would reply to, or auto-send about, the wrong thread.
    """
    deadline = time.time() + timeout
    controls_seen = False
    while time.time() < deadline:
        has_controls = _find_first(page, [
            "button[aria-label='Reply all']",
            "button[aria-label='Reply']",
            "[aria-label='Forward']",
            "button[aria-label='More options']",
            "button[aria-label='More actions']",
        ]) is not None or _find_by_text(page, "Forward") is not None
        if has_controls:
            controls_seen = True
            if num is None:
                return True
            pane = reading_pane(page)
            try:
                text = pane.text if pane is not None else ""
            except Exception:
                text = ""
            if text and label_matches_case(text, num):
                return True
        time.sleep(0.5)
    if controls_seen and num is not None:
        log.warning("reading pane never showed case %s - not acting on it", num)
    return False


def compose_is_open(page):
    el = _find_first(page, [
        "button[aria-label='Send']",
        "button[aria-label^='Send (']",
        "button[title^='Send']",
    ])
    return el is not None


def _ensure_mailbox(page, opts, ui):
    """Make sure the mailbox (search box) is reachable. If the session was
    signed out, pause and wait for the analyst to sign back in."""
    if get_search_box(page) is not None:
        return True
    try:
        page.get(opts.url)
    except Exception:
        pass
    if wait_for_mailbox(page, 15) is not None:
        return True
    log.warning("mailbox not reachable - waiting for the analyst to sign in")
    try:
        log.info("  current url: %s", page.current_url)
    except Exception:
        pass
    ui.log("    Signed out? Sign in again in the automated browser window "
           "(choose 'Yes' at 'Stay signed in?'). It waits for the inbox.")
    ok = wait_for_mailbox(page, 1800, ui) is not None
    log.info("  signed back in: %s", ok)
    return ok


def _click_folder(page, name):
    """Click a folder (e.g. 'Sent Items', 'Inbox') in the OWA left nav."""
    el = _find_first(page, [
        f"[role='treeitem'][title='{name}']",
        f"[title='{name}']",
    ])
    if el is None:
        try:
            els = page.find_elements(By.XPATH, f"//span[text()='{name}']")
            el = next((e for e in els if e.is_displayed()), None)
        except Exception:
            el = None
    if el is not None:
        try:
            el.click()
            time.sleep(1.5)
            return True
        except Exception:
            return False
    return False


def verify_sent_web(page, num, ui, since=None):
    """After the draft closed, look for the reply in Sent Items.

    `since` is when this send happened: a reply older than that belongs to
    an earlier run, and counting it would confirm a send that never left.
    True = found, False = definitely not there, None = couldn't check.
    """
    if not _click_folder(page, "Sent Items"):
        return None
    try:
        # A just-sent reply sits in Outbox briefly and the folder list lags
        # behind it. Being impatient here reports a perfectly good send as
        # unverified, which sends the analyst chasing nothing - so give it
        # a proper window and re-open the folder half way through.
        for attempt in range(8):
            # must use the scoped lookup: the raw listbox selector also
            # matches leftover search suggestions, which would "verify" a
            # reply that was never sent
            for it in _visible_results(page)[:12]:
                label = label_of(it)
                if not label_matches_case(label, num):
                    continue
                if since is not None:
                    stamp = parse_row_time(label)
                    if stamp is not None and stamp < since - timedelta(minutes=5):
                        log.debug("ignoring an older reply for %s (%s)",
                                  num, stamp)
                        continue        # a reply from a previous run
                log.debug("sent-items hit for %s on attempt %d",
                          num, attempt + 1)
                return True
            if attempt == 3:
                _click_folder(page, "Inbox")     # force the list to refresh
                time.sleep(0.6)
                _click_folder(page, "Sent Items")
            time.sleep(1.5)
        return False
    except Exception:
        return None
    finally:
        _click_folder(page, "Inbox")


def make_sent_watcher(page):
    """True once a compose pane was seen open and has since closed (sent or
    discarded). Two consecutive 'closed' polls guard against UI re-renders."""
    state = {"seen_open": False, "closed": 0}

    def watch():
        if compose_is_open(page):
            state["seen_open"] = True
            state["closed"] = 0
            return False
        if not state["seen_open"]:
            return False
        state["closed"] += 1
        return state["closed"] >= 2

    return watch


def screenshot(page, shots_dir, name):
    try:
        shots_dir.mkdir(exist_ok=True)
        path = shots_dir / f"{name}_{datetime.now():%H%M%S}.png"
        page.save_screenshot(str(path))
        return path
    except Exception:
        return None


# Row colours in the review sheet: done vs still needs a human.
STATUS_STYLE = {
    "SENT":       ("Replied",        "C6EFCE"),   # green
    "SENT_UNVERIFIED": ("Replied (unconfirmed)", "FFEB9C"),   # amber
    "SKIPPED":    ("Skipped",        "FFEB9C"),   # amber
    "NOT_FOUND":  ("NOT FOUND",      "FFC7CE"),   # red
    "ERROR":      ("ERROR",          "FFC7CE"),   # red
    "QUIT":       ("Stopped",        "E7E6E6"),   # grey
    "DUPLICATE":  ("Duplicate row",  "E7E6E6"),   # grey
    "PENDING":    ("Not done yet",   "FFF2CC"),   # pale yellow
}


def write_status_xlsx(src_path, column, status_by_case, out_path,
                      duplicate_lines=None, reason_by_case=None):
    """Copy the input sheet, add a status column and colour every row:
    green = replied, red = still needs a human, yellow = not done yet.

    duplicate_lines marks rows repeating a case handled further up, so a
    second row for the same case is not painted as though it were replied
    to in its own right.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        log.warning("openpyxl missing - no review sheet written")
        return None
    duplicate_lines = duplicate_lines or set()
    reason_by_case = reason_by_case or {}
    stats = {}
    headers, rows = read_csv_rows(src_path, stats)
    line_numbers = stats.get("line_numbers") or list(range(2, len(rows) + 2))
    wb = Workbook()
    ws = wb.active
    ws.title = "follow-up"
    ws.append(list(headers) + ["Follow-up status", "Why"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row, line in zip(rows, line_numbers):
        ws.append([row.get(h, "") for h in headers])
        m = CASE_NUM_RE.search(str(row.get(column) or ""))
        if line in duplicate_lines:
            status = "DUPLICATE"
        else:
            status = status_by_case.get(m.group(0), "PENDING") if m else "PENDING"
        text, colour = STATUS_STYLE.get(status, STATUS_STYLE["PENDING"])
        ws.cell(row=ws.max_row, column=len(headers) + 1, value=text)
        if status == "DUPLICATE":
            why = "Same case appears earlier in the sheet - handled there"
        elif status == "PENDING":
            why = "The run ended before reaching this case"
        else:
            why = reason_by_case.get(m.group(0), "") if m else ""
        ws.cell(row=ws.max_row, column=len(headers) + 2, value=why)
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill

    widths = [max(10, min(44, len(str(h)) + 4)) for h in headers] + [22, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers) + 2).column_letter}" \
                         f"{ws.max_row}"
    wb.save(out_path)
    return out_path

# ------------------------------------------------------------------ main flow

def process_case(page, num, opts, shots_dir, ui):
    """Returns (status, detail). Statuses: SENT, SKIPPED, NOT_FOUND, QUIT, ERROR."""
    if not _ensure_mailbox(page, opts, ui):
        return "ERROR", "mailbox not reachable - sign in, then use Retry"

    queries = build_queries(num)
    items, used_query = None, None
    for q in queries:
        try:
            items = run_search(page, q, opts.settle, num=num)
        except RuntimeError:
            if not _ensure_mailbox(page, opts, ui):
                return ("ERROR", "signed out - sign in, then use Retry",
                        "Outlook signed out part way through this case")
            items = run_search(page, q, opts.settle, num=num)
        if items:
            used_query = q
            break
        ui.log(f"    no match for: {q}")

    if not items:
        shot = screenshot(page, shots_dir, f"case_{num}_notfound")
        log.warning("case %s NOT FOUND after %d searches; screenshot: %s",
                    num, len(queries), shot)
        try:
            log.debug("  page url at failure: %s", page.current_url)
        except Exception:
            pass
        tried = ", ".join(queries[:3])
        return ("NOT_FOUND",
                f"no mail matched (tried {len(queries)} searches)",
                f"Searched {len(queries)} ways ({tried}, ...) - no mail in "
                f"this mailbox mentions {num}")

    try:
        count = len(items)
        ordered = sort_newest_first(items)
        first = ordered[0]
        label = label_of(first)[:120]
        if count > 1:
            ui.log(f"    {count} results - replying to the most recent one.")
        log.info("opening: %s", label)
        first.click()
        time.sleep(1.0)
        # must be showing THIS case before anything is read or clicked
        if not wait_for_reading_pane(page, num=num):
            shot = screenshot(page, shots_dir, f"case_{num}_not_opened")
            log.error("reading pane did not show case %s; screenshot: %s",
                      num, shot)
            return ("ERROR", f"mail for {num} did not open",
                    "Found the mail but the reading pane never showed "
                    f"case {num} - nothing was clicked")
    except Exception as e:
        shot = screenshot(page, shots_dir, f"case_{num}_open_fail")
        log_exception(f"case {num}: opening the result (screenshot {shot})", e)
        return ("ERROR", f"could not open result: {e}",
                f"Could not open the matching mail: {str(e)[:100]}")

    # decide about auto-send BEFORE replying, while the original mail is
    # what's on screen
    auto_ok, auto_why = False, "auto-send off"
    if getattr(opts, "auto_send", False):
        last_header, body = read_last_message(page)
        log.debug("latest message header: %s", last_header[:200] or "(none)")
        auto_ok, auto_why = should_auto_send(
            last_header, body, getattr(opts, "auto_keyword", "follow up"),
            getattr(opts, "auto_sender", ""))
        log.info("auto-send check for %s: %s (%s)", num,
                 "SEND" if auto_ok else "manual", auto_why)

    if not click_reply_all(page):
        shot = screenshot(page, shots_dir, f"case_{num}_replyall_fail")
        log.error("Reply all not found (toolbar or ... menu); screenshot: %s",
                  shot)
        return ("ERROR", "Reply all button not found",
                "Mail opened but Reply all was not there, in the toolbar "
                "or the ... menu - reply to this one by hand")
    log.info("reply-all draft opened; compose detected: %s",
             compose_is_open(page))

    if auto_ok:
        time.sleep(1.5)                      # let the draft finish building
        ui.log(f"    Auto-send: {auto_why}")
        if not click_send(page, num=num):
            shot = screenshot(page, shots_dir, f"case_{num}_send_fail")
            log.error("auto-send did not send; screenshot: %s - leaving the "
                      "draft for review", shot)
            ui.log("    Didn't send automatically - review this one by hand.")
        else:
            sent_at = datetime.now()
            time.sleep(2.5)
            verified = verify_sent_web(page, num, ui, since=sent_at)
            log.info("auto-sent %s; sent-items check: %s", num, verified)
            if verified is False:
                ui.log("    !!! Sent, but not found in Sent Items - check it.")
                return ("ERROR", label + " (auto-sent, unconfirmed)",
                        f"Auto-sent ({auto_why}), but no matching reply "
                        "appeared in Sent Items - check it by hand")
            if verified is None:
                # say so rather than claiming a check that never ran
                ui.log("    Sent automatically (could not check Sent Items).")
                return ("SENT", label + " (auto-sent, not verified)",
                        f"Auto-sent ({auto_why}); the Sent Items check "
                        "could not run")
            ui.log("    Sent automatically and verified.")
            return ("SENT", label + " (auto-sent)",
                    f"Auto-sent ({auto_why}) and confirmed in Sent Items")

    time.sleep(1.0)
    if not compose_is_open(page):
        ui.log("    (couldn't confirm the draft opened - check the browser)")

    ui.log(f"    Match ({count} result(s), query {used_query}):")
    if label:
        ui.log(f"    {label}")
    manual_note = ("" if not getattr(opts, "auto_send", False)
                   else f" (auto-send did not apply: {auto_why})")
    ui.log("    Reply-all draft is open in the browser. Review, edit, hit Send.")
    _notify(ui, num, "REVIEW", label)
    choice = _ask_with_watch(
        ui,
        "    [Enter]=sent, next case | s=skip | r=retry this case | q=quit > ",
        ("", "s", "r", "q"),
        make_sent_watcher(page),
        kind="review",
    )
    if choice == "s":
        return ("SKIPPED", label,
                "Draft was opened; the analyst skipped it without "
                "sending" + manual_note)
    if choice == "r":
        return process_case(page, num, opts, shots_dir, ui)
    if choice == "q":
        return "QUIT", "", "Analyst stopped the run at this case"

    # draft closed (auto-detected) or analyst said sent - verify in Sent Items
    if choice == "auto":
        ui.log("    Draft closed - checking Sent Items...")
    else:
        ui.log("    Checking Sent Items...")
    ok = verify_sent_web(page, num, ui)
    log.info("sent-items check for %s -> %s", num,
             {True: "found", False: "not found", None: "could not check"}[ok])
    if ok is True:
        ui.log("    Verified in Sent Items.")
        return ("SENT", label,
                "Analyst sent it; reply confirmed in Sent Items"
                + manual_note)
    if ok is None:
        return ("SENT", label + " (send not verified)",
                "Analyst sent it; the Sent Items check could not run"
                + manual_note)
    ui.log("    !!! Nothing matching in Sent Items.")
    c2 = ui.ask(
        "    Draft closed but I can't find it in Sent Items - was it sent? "
        "[Enter]=yes | s=no, mark skipped | r=retry this case > ",
        ("", "s", "r"),
        kind="verify",
    )
    if c2 == "s":
        return ("SKIPPED", label + " (draft closed without sending)",
                "Draft was closed or discarded without being sent"
                + manual_note)
    if c2 == "r":
        return process_case(page, num, opts, shots_dir, ui)
    return ("SENT", label,
            "Analyst confirmed they sent it, though it was not found in "
            "Sent Items" + manual_note)


def list_monitors():
    """Attached monitors as [{index, x, y, w, h, primary}], left-to-right.

    Uses Win32 EnumDisplayMonitors so no extra package is needed. Returns []
    if the call isn't available (non-Windows, or an odd display driver).
    """
    try:
        import ctypes
        from ctypes import wintypes
    except Exception:
        return []
    found = []
    try:
        proc = ctypes.WINFUNCTYPE(
            ctypes.c_int, ctypes.c_ulong, ctypes.c_ulong,
            ctypes.POINTER(wintypes.RECT), wintypes.LPARAM)

        def _cb(_hmon, _hdc, lprect, _data):
            r = lprect.contents
            found.append({"x": r.left, "y": r.top,
                          "w": r.right - r.left, "h": r.bottom - r.top})
            return 1

        ctypes.windll.user32.EnumDisplayMonitors(0, 0, proc(_cb), 0)
    except Exception:
        return []
    found.sort(key=lambda m: (m["x"], m["y"]))
    for i, m in enumerate(found, 1):
        m["index"] = i
        m["primary"] = (m["x"] == 0 and m["y"] == 0)
    return found


def describe_monitors():
    mons = list_monitors()
    return [f"Monitor {m['index']}: {m['w']}x{m['h']}"
            + (" (primary)" if m["primary"] else "") for m in mons]


def place_window(driver, choice="largest", ui=None):
    """Move the browser to the chosen monitor and maximise it there.

    Full width matters: on a narrow window OWA hides Reply all behind the
    '...' menu, so analysts shouldn't have to resize anything themselves.
    choice: 'largest' | 'primary' | monitor number (1-based) | 'current'
    """
    mons = list_monitors()
    target = None
    if mons and str(choice) != "current":
        if str(choice).isdigit():
            target = next((m for m in mons if m["index"] == int(choice)), None)
        elif choice == "primary":
            target = next((m for m in mons if m["primary"]), None)
        if target is None:  # 'largest' or a bad choice
            # equal-sized screens are common; put it on the primary one,
            # which is where the analyst is actually looking
            target = max(mons, key=lambda m: (m["w"] * m["h"], m["primary"]))
    try:
        if target:
            # nudge inside the target monitor first, so maximise picks it
            driver.set_window_rect(x=target["x"] + 40, y=target["y"] + 40,
                                   width=min(1400, target["w"] - 80),
                                   height=min(900, target["h"] - 120))
            time.sleep(0.4)
        driver.maximize_window()
        time.sleep(0.4)
        size = driver.get_window_size()
        if target and size.get("width", 0) < target["w"] * 0.8:
            # maximise didn't take - set the rect explicitly
            driver.set_window_rect(x=target["x"], y=target["y"],
                                   width=target["w"], height=target["h"] - 40)
            size = driver.get_window_size()
        if ui:
            where = f" on monitor {target['index']}" if target else ""
            ui.log(f"Browser window: {size.get('width')}x{size.get('height')}{where}")
        return size
    except Exception as e:
        if ui:
            ui.log(f"(couldn't position the window: {e})")
        return None


def _release_stale_profile(profile_dir, ui=None):
    """Close leftover Chrome/driver processes still holding OUR profile.

    Chrome refuses to reuse a profile another process owns and quietly
    starts a blank one instead - which looks like being signed out. Only
    processes whose command line names this profile directory are touched,
    so the analyst's own Chrome is never disturbed.
    """
    try:
        import psutil
    except ImportError:
        return 0
    target = str(Path(profile_dir).resolve()).lower()
    victims = []
    for proc in psutil.process_iter(["name", "cmdline"]):
        name = (proc.info.get("name") or "").lower()
        if not any(n in name for n in ("chrome", "chromedriver", "uc_driver")):
            continue
        cmdline = " ".join(proc.info.get("cmdline") or []).lower()
        if target and target in cmdline:
            victims.append(proc)
    if not victims:
        return 0
    if ui:
        ui.log(f"Closing {len(victims)} leftover browser process(es) from a "
               "previous run...")
    for proc in victims:
        try:
            proc.terminate()
        except Exception:
            pass
    gone, alive = psutil.wait_procs(victims, timeout=8)
    for proc in alive:
        try:
            proc.kill()
        except Exception:
            pass
    lock = Path(profile_dir) / "lockfile"
    try:
        if lock.exists():
            lock.unlink()
    except Exception:
        pass
    time.sleep(1.0)  # let Chrome flush the cookie DB before relaunching
    return len(victims)


def chrome_version():
    """Installed Chrome version, e.g. '151.0.7922.76' (None if not found)."""
    try:
        import winreg
        for root in (winreg.HKEY_CURRENT_USER, winreg.HKEY_LOCAL_MACHINE):
            try:
                with winreg.OpenKey(root, r"Software\Google\Chrome\BLBeacon") as k:
                    return winreg.QueryValueEx(k, "version")[0]
            except OSError:
                continue
    except Exception:
        pass
    for path in (r"C:\Program Files\Google\Chrome\Application",
                 r"C:\Program Files (x86)\Google\Chrome\Application"):
        try:
            versions = [d.name for d in Path(path).iterdir()
                        if d.is_dir() and d.name[0].isdigit()]
            if versions:
                return sorted(versions)[-1]
        except Exception:
            continue
    return None


def find_local_driver(version=None):
    """A chromedriver.exe shipped with the tool, preferring one whose major
    version matches Chrome. Lets the tool work where driver downloads are
    blocked."""
    major = (version or "").split(".")[0]
    beside = SCRIPT_DIR / "chromedriver.exe"
    if beside.exists():
        return beside
    drivers = SCRIPT_DIR / "drivers"
    if not drivers.is_dir():
        return None
    candidates = sorted(drivers.glob("chromedriver*.exe"))
    if major:
        for path in candidates:
            if major in path.stem:
                return path
    return candidates[0] if candidates else None


def driver_info():
    """What Chrome is installed, and whether we have a driver to match."""
    version = chrome_version()
    major = (version or "?").split(".")[0]
    out = [f"Chrome installed : {version or 'not detected'}"]
    local = find_local_driver(version)
    if local:
        matched = major in local.stem or local.name == "chromedriver.exe"
        out.append(f"Bundled driver   : {local.name}"
                   + ("" if matched else "  (MAJOR VERSION MISMATCH)"))
    else:
        out.append("Bundled driver   : none - Selenium will download one, "
                   "which needs internet")
    if version and (not local or major not in local.stem):
        out += [
            "",
            f"To work offline, download the chromedriver for Chrome {major}:",
            "  https://googlechromelabs.github.io/chrome-for-testing/",
            f"  pick win64 for version {major}.x, then put chromedriver.exe",
            f"  in: {SCRIPT_DIR}",
        ]
    return out


def _launch_browser(opts, ui=None):
    """Start Chrome with a persistent profile.

    Plain Selenium WebDriver by default: ordinary, visible browser
    automation, which is what gets signed off in a managed environment.
    engine='stealth' swaps in SeleniumBase's undetected mode - only needed
    for personal outlook.com accounts, whose risk engine revokes sessions it
    can tell are automated.
    """
    profile_dir = Path(opts.profile_dir)
    profile_dir.mkdir(parents=True, exist_ok=True)
    _release_stale_profile(profile_dir, ui)
    engine = getattr(opts, "engine", "standard")

    if engine == "stealth":
        try:
            from seleniumbase import Driver
        except ImportError:
            raise RuntimeError("stealth mode needs seleniumbase - run: "
                               "pip install seleniumbase (not needed for "
                               "normal use)")
        driver = Driver(uc=True, headless=False, user_data_dir=str(profile_dir))
    else:
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
        except ImportError:
            raise RuntimeError("selenium is not installed - run: "
                               "pip install selenium")
        options = Options()
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument("--start-maximized")
        # Prefer a driver shipped with the tool - environments that block
        # driver downloads are exactly where this is used.
        local_driver = find_local_driver(chrome_version())
        if local_driver:
            log.info("using bundled driver: %s", local_driver.name)
            service = Service(executable_path=str(local_driver))
        else:
            log.info("no bundled driver - Selenium will fetch one")
            service = Service()
        try:
            driver = webdriver.Chrome(options=options, service=service)
        except Exception as e:
            log_exception("Chrome failed to start", e)
            raise RuntimeError(
                f"{e}\n\n" + "\n".join(driver_info())) from None

    if ui:
        ui.log("Browser: Chrome ("
               + ("undetected mode" if engine == "stealth"
                  else "standard Selenium") + ")")
    place_window(driver, getattr(opts, "monitor", "largest"), ui)
    return driver


def _case_loop(cases, opts, ui, run_one, on_error=None, results=None):
    """Shared per-case loop: bookkeeping, not-found pauses, quit handling.

    `results` may be a caller-owned list so finished cases survive even if
    the run blows up part way through.
    """
    results = results if results is not None else []
    for i, num in enumerate(cases, 1):
        # auto-send can carry a whole run without ever showing a prompt, so
        # check for a stop request between cases or Stop would be unusable
        stop = getattr(ui, "stop_requested", None)
        if callable(stop) and stop():
            ui.log("    Stopping at user request.")
            log.info("run stopped by the analyst after %d case(s)", len(results))
            break
        ui.log(f"[{i}/{len(cases)}] SOC{num}")
        _notify(ui, num, "WORKING")
        reason = ""
        try:
            outcome = run_one(num)
            if len(outcome) == 3:
                status, detail, reason = outcome
            else:
                status, detail = outcome
        except KeyboardInterrupt:
            status, detail = "QUIT", "interrupted"
            reason = "Run interrupted"
        except Exception as e:
            if on_error:
                on_error(num)
            log_exception(f"case {num} failed", e)
            status, detail = "ERROR", str(e)[:200]
            reason = f"Something went wrong: {str(e)[:120]}"
        _notify(ui, num, status, detail)
        results.append({"case": f"SOC{num}", "status": status,
                        "detail": detail, "reason": reason})
        if status == "NOT_FOUND":
            ui.log(f"    !!! NOT FOUND - flagged: SOC{num}")
            if not opts.no_pause:
                if ui.ask("    [Enter]=continue | q=quit > ", ("", "q"),
                          kind="notfound") == "q":
                    break
        elif status == "ERROR":
            ui.log(f"    !!! ERROR: {detail}")
        elif status == "QUIT":
            ui.log("    Stopping at user request.")
            break
        if status == "SENT" and i < len(cases):
            delay = getattr(opts, "send_delay", 5.0)
            if delay > 0:
                ui.log(f"    Waiting {delay:g}s before the next case...")
                time.sleep(delay)
    return results


def _run_web(cases, opts, ui, results=None):
    shots_dir = SCRIPT_DIR / "screenshots"
    log.info("run starting: %d case(s), url=%s, engine=%s, monitor=%s, "
             "settle=%ss, send_delay=%ss, pause_on_notfound=%s",
             len(cases), opts.url, getattr(opts, "engine", "standard"),
             getattr(opts, "monitor", "largest"), opts.settle,
             getattr(opts, "send_delay", 0), not opts.no_pause)
    log.debug("cases: %s", cases)
    driver = _launch_browser(opts, ui)
    log_browser(driver)
    try:
        driver.get(opts.url)
        ui.log("If a login page appears, sign in manually (MFA included) - "
               "it waits until the inbox is open.")
        if wait_for_mailbox(driver, 1800, ui) is None:
            raise RuntimeError("the inbox never opened - sign in, then run again")
        ui.log("Mailbox ready.\n")

        results = _case_loop(
            cases, opts, ui,
            run_one=lambda num: process_case(driver, num, opts, shots_dir, ui),
            on_error=lambda num: screenshot(driver, shots_dir, f"case_{num}_error"),
            results=results,
        )
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        time.sleep(2)  # let Chrome write out cookies before sweeping orphans
        _release_stale_profile(opts.profile_dir)
    return results


def sign_in_only(opts, ui):
    """Open the mailbox just so the analyst can sign in once; the session is
    saved to the profile, then the browser closes cleanly."""
    driver = _launch_browser(opts, ui)
    try:
        driver.get(opts.url)
        ui.log("Sign in to Outlook in the browser window "
               "(choose 'Yes' at 'Stay signed in?').")
        ui.log("Waiting until the inbox opens - take as long as you need.")
        if wait_for_mailbox(driver, 1800, ui) is None:
            ui.log("Gave up waiting - the inbox never opened.")
            return False
        ui.log("Signed in. Saving the session...")
        time.sleep(3)  # give Chrome a moment to persist the login cookies
        return True
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        time.sleep(2)
        _release_stale_profile(opts.profile_dir)


def run_followups(src_path, column, cases, opts, ui):
    """Run every case, then write the result files. Returns a summary."""
    setup_logging()          # no-op if the caller already started logging
    log.info("input sheet   : %s (column %r)", src_path, column)

    # An hour of a night shift can sit in these results, so a failure after
    # the loop - a browser dying, the profile vanishing - must not throw
    # them away. Whatever was finished still gets written and reported.
    collected = []
    run_error = None
    try:
        collected = _run_web(cases, opts, ui, collected)
    except Exception as e:
        run_error = e
        log_exception("run stopped early", e)
        ui.log(f"    !!! Run stopped: {e}")
        ui.log(f"    Keeping the {len(collected)} case(s) already done.")
    results = collected


    stamp = f"{datetime.now():%Y%m%d_%H%M%S}"
    out_csv = SCRIPT_DIR / f"followup_results_{stamp}.csv"
    with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["case", "status", "reason", "detail"],
                           extrasaction="ignore")
        w.writeheader()
        for r in results:
            w.writerow({"case": r["case"], "status": r["status"],
                        "reason": r.get("reason", ""), "detail": r["detail"]})

    # anything not replied to needs a human: not-found AND errored cases
    flagged = [r["case"] for r in results
               if r["status"] in ("NOT_FOUND", "ERROR")]
    not_found = [r["case"] for r in results if r["status"] == "NOT_FOUND"]
    errored = [r["case"] for r in results if r["status"] == "ERROR"]
    done = sum(1 for r in results if r["status"] == "SENT")
    ui.log(f"\nDone. {done} replied, {len(not_found)} not found, "
           f"{len(errored)} errored, {len(results)} processed. "
           f"Log: {out_csv.name}")
    if flagged:
        ui.log("NEEDS MANUAL FOLLOW-UP:")
        for r in results:
            if r["status"] in ("NOT_FOUND", "ERROR"):
                ui.log(f"  {r['case']} ({r['status']})")

    # Colour-coded copy of the analyst's own sheet: green = replied,
    # red = still needs a human, yellow = never got to it.
    xlsx = SCRIPT_DIR / f"followup_review_{stamp}.xlsx"
    status_by_case, reason_by_case = {}, {}
    for r in results:
        num = r["case"].replace("SOC", "")
        status = r["status"]
        if status == "SENT" and "not verified" in r["detail"]:
            status = "SENT_UNVERIFIED"      # replied, but we could not confirm
        status_by_case[num] = status
        reason_by_case[num] = r.get("reason", "")
    dup_stats = {}
    try:
        extract_cases(src_path, column, dup_stats)
    except Exception:
        pass
    try:
        if write_status_xlsx(src_path, column, status_by_case, xlsx,
                             dup_stats.get("duplicate_lines"), reason_by_case):
            ui.log(f"Review sheet (green = replied, red = needs follow-up): "
                   f"{xlsx.name}")
        else:
            xlsx = None
    except Exception as e:
        # never lose the run's results just because the copy failed
        ui.log(f"(couldn't write the review sheet: {e})")
        xlsx = None
    log.info("run finished: %d replied, %d not found, %d errored, %d processed",
             done, len(not_found), len(errored), len(results))
    if LOG_PATH:
        ui.log(f"Log (send this if something went wrong): {LOG_PATH.name}")
    return {"results": results, "out_csv": out_csv, "flagged": flagged,
            "not_found": not_found, "errors": errored, "xlsx": xlsx,
            "log": LOG_PATH}


def collect_diagnostics(out_path=None):
    """Zip the logs, screenshots and result files so they can be sent on.

    Contains case numbers and email subjects, so treat it like the XSOAR
    export itself. The signed-in browser profile is never included.
    """
    import zipfile
    out_path = Path(out_path or
                    SCRIPT_DIR / f"diagnostics_{datetime.now():%Y%m%d_%H%M%S}.zip")
    added = []
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as z:
        for folder in ("logs", "screenshots"):
            base = SCRIPT_DIR / folder
            if base.is_dir():
                # newest first: sorting by name could keep last month's
                # screenshots and drop the ones from the run that just failed
                files = sorted((f for f in base.iterdir() if f.is_file()),
                               key=lambda f: f.stat().st_mtime, reverse=True)
                for f in files[:40]:
                    z.write(f, arcname=f"{folder}/{f.name}")
                    added.append(f"{folder}/{f.name}")
        for f in sorted(SCRIPT_DIR.glob("followup_results_*.csv"))[-5:]:
            z.write(f, arcname=f.name)
            added.append(f.name)
        env = "\n".join(environment_report())
        z.writestr("environment.txt", env)
        added.append("environment.txt")
    return out_path, added

# ------------------------------------------------------------------------ CLI

def main():
    ap = argparse.ArgumentParser(description="SOC nightshift follow-up helper")
    # not required: --driver-info, --list-monitors and --diagnostics are
    # useful on their own, without a sheet to point at
    ap.add_argument("--csv", help="XSOAR export (.csv or .xlsx)")
    ap.add_argument("--column", help="column holding case numbers (auto-detected if omitted)")
    ap.add_argument("--url", default="https://outlook.office.com/mail/",
                    help="Outlook web URL (use https://outlook.live.com/mail/ for a personal test mailbox)")
    ap.add_argument("--profile-dir", default=str(SCRIPT_DIR / "uc_profile"),
                    help="Browser profile folder (keeps you logged in between runs)")
    ap.add_argument("--settle", type=float, default=3.0,
                    help="Seconds to wait after searching before reading results")
    ap.add_argument("--send-delay", type=float, default=5.0,
                    help="Seconds to pause after a sent reply before the next case")
    ap.add_argument("--auto-send", action="store_true",
                    help="Send the reply automatically when the open mail "
                         "contains --auto-keyword AND came from --auto-sender. "
                         "Anything else still waits for you.")
    ap.add_argument("--auto-sender", default="",
                    help="Sender that permits auto-send - email address or "
                         "display name, e.g. alerts@example.com or 'SOC Alerts'")
    ap.add_argument("--auto-keyword", default="follow up",
                    help="Phrase that must appear in the mail (default: 'follow up')")
    ap.add_argument("--engine", default="standard", choices=["standard", "stealth"],
                    help="'standard' = plain Selenium Chrome automation (default). "
                         "'stealth' = SeleniumBase undetected mode, only for "
                         "personal outlook.com accounts that keep signing out")
    ap.add_argument("--monitor", default="largest",
                    help="Which screen to run the browser on: largest (default), "
                         "primary, current, or a monitor number")
    ap.add_argument("--list-monitors", action="store_true",
                    help="Show the detected monitors and exit")
    ap.add_argument("--driver-info", action="store_true",
                    help="Show the Chrome version and which driver will be "
                         "used, then exit")
    ap.add_argument("--limit", type=int, help="Only process the first N cases (for testing)")
    ap.add_argument("--no-pause", action="store_true",
                    help="Don't pause on not-found cases, just flag and continue")
    ap.add_argument("--parse-only", action="store_true",
                    help="Just show the case numbers read from the file, then exit")
    ap.add_argument("--verbose", action="store_true",
                    help="Also print the detailed log to the console")
    ap.add_argument("--diagnostics", action="store_true",
                    help="Zip up logs, screenshots and result files to send on, "
                         "then exit")
    ap.add_argument("--check", action="store_true",
                    help="Dry-run a new export: what would be processed and what "
                         "would be ignored. Touches nothing.")
    args = ap.parse_args()
    setup_logging(verbose=args.verbose)

    if args.diagnostics:
        path, added = collect_diagnostics()
        print(f"Diagnostics bundle: {path}")
        for name in added:
            print(f"   {name}")
        print("\nSend this on if something went wrong. It holds case numbers "
              "and email\nsubjects, so treat it like the XSOAR export. Your "
              "signed-in browser\nsession is NOT included.")
        return

    if args.driver_info:
        print("\n".join(driver_info()))
        return

    if args.list_monitors:
        mons = describe_monitors()
        print("\n".join(mons) if mons else "No monitors detected.")
        return

    if not args.csv:
        sys.exit("ERROR: --csv is needed to run cases. For a quick check try:\n"
                 "  python soc_followup.py --driver-info\n"
                 "  python soc_followup.py --csv <your export> --check")

    if args.check:
        try:
            print(format_preflight(preflight(args.csv, args.column)))
        except ValueError as e:
            sys.exit(f"ERROR: {e}")
        return

    stats = {}
    try:
        column, cases = extract_cases(args.csv, args.column, stats)
    except ValueError as e:
        sys.exit(f"ERROR: {e}")
    print(f"File: {args.csv}")
    hidden = stats.get("hidden_skipped", 0)
    hidden_note = f" ({hidden} filtered-out row(s) skipped)" if hidden else ""
    print(f"Case-number column: '{column}' -> {len(cases)} unique case(s){hidden_note}")
    if args.limit:
        cases = cases[: args.limit]
        print(f"Limited to first {len(cases)} case(s)")
    if args.parse_only:
        for n in cases:
            print(f"  SOC{n}")
        return
    if not cases:
        sys.exit("Nothing to do - no case numbers found.")

    try:
        run_followups(args.csv, column, cases, args, ConsoleUI())
    except RuntimeError as e:
        sys.exit(f"ERROR: {e}")


if __name__ == "__main__":
    main()
