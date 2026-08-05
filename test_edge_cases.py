#!/usr/bin/env python3
"""Edge-case tests for the SOC follow-up tool.

Complements audit_simulation.py (which covers the happy paths and the main
failure modes). This one pokes at the nasty stuff: lookalike case numbers,
analyst stopping mid-run, retries, a mailbox that signs out mid-run, Sent
Items being unreadable, odd spreadsheets.

    python test_edge_cases.py
"""

import sys
import tempfile
import time
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parent))
import audit_simulation as A
import soc_followup as core

PASS, FAIL = [], []


def check(label, ok, note=""):
    (PASS if ok else FAIL).append(label)
    print(f"[{'OK ' if ok else 'FAIL'}] {label}" + (f"  ({note})" if note else ""))


def sheet_from(rows, folder, hide_last=False, header="id"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append([header, "name"])
    for r in rows:
        ws.append([r, "alert"])
    if hide_last and ws.max_row > 1:
        ws.row_dimensions[ws.max_row].hidden = True
    path = Path(folder) / "edge.xlsx"
    wb.save(path)
    return path


class UI:
    """Scripted analyst; `answers` maps case number -> what they do."""

    def __init__(self, app, answers=None, watch_answers=None):
        self.app, self.answers = app, answers or {}
        self.watch_answers = watch_answers or {}
        self.lines, self.updates, self.current = [], [], ""

    def log(self, m):
        self.lines.append(str(m))

    def case_update(self, num, status, detail=""):
        self.updates.append((str(num), status))
        if status == "WORKING":
            self.current = str(num)

    def ask(self, prompt, choices, kind=None):
        return self.answers.get(self.current, {}).get(kind or "ask", "")

    def ask_or_watch(self, prompt, choices, watch_fn, poll_s=0.1, kind=None):
        forced = self.watch_answers.get(self.current)
        if forced is not None:
            return forced
        for i in range(60):
            if i == 3:
                self.app.analyst_sends()
            if watch_fn():
                return "auto"
            time.sleep(0.05)
        return ""


def run(cases, app, ui, **opt):
    tmp = Path(tempfile.mkdtemp(prefix="edge_"))
    core.SCRIPT_DIR = tmp
    core._launch_browser = lambda opts, ui=None: app
    core._release_stale_profile = lambda *a, **k: 0
    opts = SimpleNamespace(backend="web", url="x", profile_dir=str(tmp / "p"),
                           settle=0.02, send_delay=0, monitor="current",
                           no_pause=True, **opt)
    sheet = sheet_from(cases, tmp)
    return core.run_followups(sheet, "id", cases, opts, ui)


print("=" * 72)
print("EDGE-CASE TESTS")
print("=" * 72)

# 1. A longer number that merely CONTAINS the case number must not match.
inbox = [A.Mail("1610529", "SOC1610529 different case entirely", "7:34 PM")]
app = A.FakeOutlook(inbox, {})
ui = UI(app)
res = run(["610529"], app, ui)
check("lookalike case number 1610529 not treated as 610529",
      res["results"][0]["status"] == "NOT_FOUND",
      res["results"][0]["status"])

# 2. ...but the real one still matches when both are present.
inbox = [A.Mail("1610529", "SOC1610529 wrong one", "6:00 PM"),
         A.Mail("610529", "SOC610529 correct one", "7:34 PM")]
app = A.FakeOutlook(inbox, {})
ui = UI(app)
res = run(["610529"], app, ui)
replied = [s for _, s in app.replies_started]
check("picks the exact case when a lookalike sits beside it",
      replied == ["SOC610529 correct one"], str(replied))

# 3. Analyst hits Stop mid-run: remaining cases untouched, results still saved.
inbox = [A.Mail("610321", "SOC610321 one", "7:00 PM"),
         A.Mail("610322", "SOC610322 two", "7:10 PM")]
app = A.FakeOutlook(inbox, {})
ui = UI(app, watch_answers={"610321": "q"})
res = run(["610321", "610322"], app, ui)
statuses = [r["status"] for r in res["results"]]
check("Stop halts the run immediately", statuses == ["QUIT"], str(statuses))
check("results file still written after Stop", Path(res["out_csv"]).exists())
# the current case's draft is already open when Stop is pressed - what must
# not happen is the NEXT case being opened
check("no later case touched after Stop",
      [n for n, _ in app.replies_started] == ["610321"],
      str(app.replies_started))

# 4. Retry re-runs the same case and can then succeed.
app = A.FakeOutlook([A.Mail("610321", "SOC610321 one", "7:00 PM")], {})


class RetryUI(UI):
    def __init__(self, app):
        super().__init__(app)
        self.first = True

    def ask_or_watch(self, prompt, choices, watch_fn, poll_s=0.1, kind=None):
        if self.first:                      # analyst retries once
            self.first = False
            return "r"
        return super().ask_or_watch(prompt, choices, watch_fn, poll_s, kind)


ui = RetryUI(app)
res = run(["610321"], app, ui)
check("Retry re-opens the case and can then send",
      res["results"][0]["status"] == "SENT" and len(app.replies_started) == 2,
      f"{res['results'][0]['status']}, drafts={len(app.replies_started)}")

# 5. Sent Items unreadable -> marked sent but explicitly "not verified".
class NoSentFolder(A.FakeOutlook):
    def find_elements(self, by, sel):
        if "Sent Items" in sel:
            return []                        # folder not clickable
        return super().find_elements(by, sel)


app = NoSentFolder([A.Mail("610321", "SOC610321 one", "7:00 PM")], {})
ui = UI(app)
res = run(["610321"], app, ui)
detail = res["results"][0]["detail"]
check("unverifiable send is labelled '(send not verified)'",
      res["results"][0]["status"] == "SENT" and "not verified" in detail,
      detail[-30:])

# 6. Signed out mid-run: the tool waits and recovers rather than failing.
class LogsOutOnce(A.FakeOutlook):
    def __init__(self, *a, **k):
        super().__init__(*a, **k)
        self.signed_out = True

    def find_elements(self, by, sel):
        if sel == "#topSearchInput" and self.signed_out:
            return []                        # login page: no search box
        return super().find_elements(by, sel)

    def get(self, url):                      # navigating restores the session
        self.signed_out = False
        return super().get(url)


app = LogsOutOnce([A.Mail("610321", "SOC610321 one", "7:00 PM")], {})
ui = UI(app)
res = run(["610321"], app, ui)
check("recovers from a mid-run sign-out instead of failing the case",
      res["results"][0]["status"] == "SENT", res["results"][0]["status"])

# 7. Spreadsheet oddities
tmp = Path(tempfile.mkdtemp(prefix="edge_sheets_"))
p = sheet_from(["610321", "610321", "610322"], tmp)
_, cases = core.extract_cases(p)
check("duplicate case numbers collapsed to one", cases == ["610321", "610322"],
      str(cases))

p = sheet_from(["12345", "12345678"], tmp)
_, cases = core.extract_cases(p)
check("5- and 8-digit case numbers both accepted",
      cases == ["12345", "12345678"], str(cases))

p = sheet_from(["610321"], tmp, hide_last=True)
stats = {}
try:
    _, cases = core.extract_cases(p, stats=stats)
    ok = cases == []
except ValueError:
    ok = True                                # also acceptable: clear error
check("sheet with every row filtered out yields no cases", ok)

p = sheet_from(["610321", "610322"], tmp, header="Case ID")
col, cases = core.extract_cases(p)
check("unusual header name still detected", col == "Case ID" and len(cases) == 2,
      f"column={col}")

from openpyxl import Workbook
wb = Workbook(); ws = wb.active
ws.append(["name", "owner"]); ws.append(["no numbers here", "someone"])
p2 = tmp / "nocol.xlsx"; wb.save(p2)
try:
    core.extract_cases(p2)
    ok, note = False, "no error raised"
except ValueError as e:
    ok, note = "auto-detect" in str(e), str(e)[:40]
check("sheet with no case column gives a clear error", ok, note)

# 8. Case number that appears only in the body preview, not the subject
inbox = [A.Mail("610321", "Weekly digest mentioning 610321 inside", "7:00 PM")]
app = A.FakeOutlook(inbox, {})
ui = UI(app)
res = run(["610321"], app, ui)
check("still replies when the number is in the row text (analyst reviews it)",
      res["results"][0]["status"] == "SENT", res["results"][0]["status"])

print("-" * 72)
print(f"{len(PASS)} passed, {len(FAIL)} failed")
if FAIL:
    for f in FAIL:
        print("   FAILED:", f)
print("=" * 72)
sys.exit(1 if FAIL else 0)
