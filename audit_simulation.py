#!/usr/bin/env python3
"""End-to-end audit of the follow-up tool against a simulated Outlook.

Runs the real code path (search -> open -> Reply all -> send detection ->
Sent Items verification -> logging -> result files) against a fake Outlook
web UI, so every branch can be checked without touching a real mailbox.

    python audit_simulation.py

Prints a pass/fail line per scenario and exits non-zero if anything failed.
"""

import csv
import shutil
import sys
import tempfile
import time
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parent))
import soc_followup as core

CSS, XPATH, TAG = "css selector", "xpath", "tag name"


# --------------------------------------------------------------- fake Outlook

class El:
    def __init__(self, label="", on_click=None, enabled=True):
        self.label, self._on_click, self.enabled = label, on_click, enabled
        self.clicks = 0

    def is_displayed(self):
        return True

    def is_enabled(self):
        return self.enabled

    def get_attribute(self, _name):
        return self.label

    @property
    def text(self):
        return self.label

    def click(self):
        self.clicks += 1
        if self._on_click:
            self._on_click()

    def send_keys(self, *_args):
        pass


class Mail:
    def __init__(self, num, subject, when):
        self.num, self.subject, self.when = num, subject, when

    @property
    def row_label(self):
        return f"Ranveer Singh {self.subject} {self.when}"


class FakeOutlook:
    """Mimics the parts of Outlook on the web the tool touches."""

    def __init__(self, inbox, behaviour):
        self.inbox = inbox                  # list[Mail]
        self.behaviour = behaviour          # per-case-number knobs
        self.sent_items = []                # list[Mail]
        self.folder = "Inbox"
        self.query = ""
        self.results = []
        self.open_mail = None
        self.menu_open = False
        self.compose_for = None
        self.replies_started = []           # (case, subject) - audit trail
        self.shots = []

    # -- analyst actions the scripted UI calls -----------------------------
    def analyst_sends(self):
        mail = self.compose_for
        self.sent_items.append(Mail(mail.num, f"RE: {mail.subject}", "now"))
        self.compose_for = None

    def analyst_discards(self):
        self.compose_for = None             # closed, nothing sent

    # -- search ------------------------------------------------------------
    def _run_query(self, q):
        self.query = q
        knob = {}
        for num, cfg in self.behaviour.items():
            if str(num) in q:
                knob = cfg
                break
        if knob.get("search_returns_inbox"):
            self.results = list(self.inbox)          # search silently ignored
            return
        if q.lower().startswith("subject:"):
            self.results = []                        # matches the live failure
            return
        term = q.strip().strip('"').lower().replace(" ", "")
        self.results = [m for m in self.inbox
                        if term and term in m.subject.lower().replace(" ", "")]

    def _rows(self):
        pool = self.sent_items if self.folder == "Sent Items" else self.results
        out = []
        for mail in pool:
            out.append(El(mail.row_label, on_click=lambda m=mail: self._open(m)))
        return out

    def _open(self, mail):
        self.open_mail = mail
        self.menu_open = False

    def _open_menu(self):
        self.menu_open = True

    def _reply_all(self):
        self.menu_open = False
        self.compose_for = self.open_mail
        self.replies_started.append((self.open_mail.num, self.open_mail.subject))

    def _switch(self, folder):
        self.folder = folder
        self.open_mail = None

    # -- WebDriver surface -------------------------------------------------
    def find_elements(self, by, sel):
        if by == TAG:                       # body text, as OWA renders it
            if self.query and not self._rows():
                return [El("We couldn't find anything to show here.")]
            return [El("Inbox")]
        if by == XPATH:
            if "'Reply all'" in sel and self.menu_open:
                return [El("Reply all", on_click=self._reply_all)]
            if "'Forward'" in sel and self.open_mail:
                return [El("Forward")]
            if "'Sent Items'" in sel:
                return [El("Sent Items", on_click=lambda: self._switch("Sent Items"))]
            if "'Inbox'" in sel:
                return [El("Inbox", on_click=lambda: self._switch("Inbox"))]
            return []

        # --- CSS ---
        if sel == "#topSearchInput":
            return [SearchBox(self)]
        # the reading pane: the tool checks this really is showing the case
        # it just clicked before acting on it
        if ("Reading pane" in sel or "ReadingPaneContainerId" in sel
                or sel == "div[role='document']"):
            if not self.open_mail:
                return []
            return [El(f"{self.open_mail.row_label} {self.open_mail.subject}")]
        if "uggest" in sel:                              # autocomplete popup
            return [El(f"{self.query} suggestion")] if self.query else []
        if sel.startswith("div[aria-label='Message list']"):
            return self._rows()
        if sel == "[role='listbox'] [role='option']":
            # generic fallback: suggestions AND rows, like the real DOM
            sugg = [El(self.query)] if self.query else []
            return sugg + self._rows()
        if "Message list" in sel or "MessageList" in sel:
            return self._rows()
        if sel.startswith("[role='treeitem'][title='") or sel.startswith("[title='"):
            name = sel.split("title='")[1].rstrip("']")
            if name in ("Sent Items", "Inbox"):
                return [El(name, on_click=lambda n=name: self._switch(n))]
            return []
        if self.open_mail and sel in (
                "button[aria-label='Reply all']", "[aria-label='Reply all']",
                "button[title='Reply all']", "[aria-label*='Reply all']"):
            knob = self.behaviour.get(self.open_mail.num, {})
            if knob.get("reply_all_hidden"):
                return []                                # narrow window
            return [El("Reply all", on_click=self._reply_all)]
        if self.open_mail and sel in ("button[aria-label='Reply']",
                                      "[aria-label='Forward']"):
            return [El("Reply")]
        if self.open_mail and sel in ("button[aria-label='More options']",
                                      "button[aria-label='More actions']"):
            return [El("More options", on_click=self._open_menu)]
        if self.compose_for and sel in ("button[aria-label='Send']",
                                        "button[aria-label^='Send (']",
                                        "button[title^='Send']"):
            return [El("Send")]
        return []

    def find_element(self, by, sel):
        found = self.find_elements(by, sel)
        if found:
            return found[0]
        if by == TAG:                                     # body text
            return El("We couldn't find anything" if not self._rows() else "mail")
        return El("")

    def get(self, _url):
        self.folder, self.open_mail, self.query = "Inbox", None, ""
        self.results = []

    def save_screenshot(self, path):
        Path(path).write_bytes(b"png")
        self.shots.append(Path(path).name)
        return True

    def quit(self):
        pass


class SearchBox(El):
    def __init__(self, app):
        super().__init__("Search")
        self.app = app
        self.buffer = ""

    def send_keys(self, *args):
        s = "".join(str(a) for a in args)
        # written as escapes on purpose: literal control characters get
        # mangled if this file is ever rewritten by a non-UTF8 tool
        if s == "\ue007":                          # Enter - run the search
            self.app._run_query(self.buffer)
        elif s in ("\ue009a", "\ue017", "\ue003"):  # ctrl+a / delete
            self.buffer = ""
        else:
            self.buffer += s
# ------------------------------------------------------------- scripted UI

class ScriptedUI:
    """Plays the analyst: sends or discards, then answers any question."""

    def __init__(self, app, plan):
        self.app, self.plan = app, plan
        self.lines, self.updates = [], []
        self.current = ""          # case being worked on right now

    def log(self, msg):
        self.lines.append(str(msg))

    def case_update(self, num, status, detail=""):
        self.updates.append((str(num), status))
        if status == "WORKING":
            self.current = str(num)

    def ask(self, prompt, choices, kind=None):
        if kind == "verify":
            return self.plan.get(self.current, {}).get("verify_answer", "")
        return ""                                        # continue

    def ask_or_watch(self, prompt, choices, watch_fn, poll_s=0.1, kind=None):
        """The draft sits open for a moment, then the analyst sends (or
        discards) it - the real watcher has to notice by itself."""
        action = self.plan.get(self.current, {}).get("action", "send")
        for i in range(60):
            if i == 3:                       # analyst acts, draft closes
                if action == "send":
                    self.app.analyst_sends()
                else:
                    self.app.analyst_discards()
            if watch_fn():
                return "auto"
            time.sleep(0.05)
        return ""                            # watcher never fired


# ------------------------------------------------------------------ scenarios

INBOX = [
    Mail("610321", "SOC610321 Suspicious login", "7:34 PM"),
    Mail("610454", "SOC 610454 test", "7:30 PM"),
    Mail("610529", "soc610529", "7:34 PM"),
    Mail("610999", "SOC610999 first alert", "Mon 9:00 AM"),
    Mail("610999", "SOC610999 latest update", "7:50 PM"),      # newest
]

BEHAVIOUR = {
    "610454": {"reply_all_hidden": True},        # only via the '...' menu
    "610888": {"search_returns_inbox": True},    # search silently fails
}

PLAN = {
    "610529": {"action": "discard", "verify_answer": "s"},
}

EXPECTED = {
    "SOC610321": ("SENT", "direct Reply all button, analyst sends"),
    "SOC610454": ("SENT", "Reply all hidden - uses the '...' menu"),
    "SOC610529": ("SKIPPED", "draft discarded - caught by Sent Items check"),
    "SOC610777": ("NOT_FOUND", "no such mail in the mailbox"),
    "SOC610888": ("NOT_FOUND", "search returned unrelated inbox rows"),
    "SOC610999": ("SENT", "two matches - replies to the newest"),
}


def build_sheet(folder):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name", "severity"])
    rows = [("610321", "Suspicious login"), ("610454", "Test alert"),
            ("610529", "Phishing report"), ("610777", "Missing case"),
            ("610888", "Search fails"), ("610999", "Duplicate mails"),
            ("123021", "Filtered out - must be skipped")]
    for cid, name in rows:
        ws.append([cid, name, "High"])
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.row_dimensions[ws.max_row].hidden = True          # the filtered-out row
    path = folder / "audit_cases.xlsx"
    wb.save(path)
    return path


def main():
    workdir = Path(tempfile.mkdtemp(prefix="soc_audit_"))
    sheet = build_sheet(workdir)
    app = FakeOutlook(INBOX, BEHAVIOUR)
    ui = ScriptedUI(app, PLAN)

    core.SCRIPT_DIR = workdir                            # keep outputs isolated
    core._launch_browser = lambda opts, ui=None: app
    core._release_stale_profile = lambda *a, **k: 0

    stats = {}
    column, cases = core.extract_cases(sheet, stats=stats)
    opts = SimpleNamespace(backend="web", url="https://outlook.live.com/mail/",
                           profile_dir=str(workdir / "p"), settle=0.05,
                           send_delay=0.2, monitor="largest", no_pause=True)

    print("=" * 72)
    print("SOC FOLLOW-UP TOOL - SIMULATED AUDIT RUN")
    print("=" * 72)
    print(f"Sheet          : {sheet.name}")
    print(f"Case column    : '{column}'")
    print(f"Cases to run   : {len(cases)}  (filter-hidden rows skipped: "
          f"{stats.get('hidden_skipped', 0)})")
    print(f"Simulated inbox: {len(INBOX)} mails\n")

    started = time.time()
    summary = core.run_followups(sheet, column, cases, opts, ui)
    elapsed = time.time() - started

    results = {r["case"]: r["status"] for r in summary["results"]}
    details = {r["case"]: r["detail"] for r in summary["results"]}
    failures = []
    if "--verbose" in sys.argv:
        for case, det in details.items():
            print(f"  detail {case}: {results[case]} - {det[:110]}")

    print("\n" + "-" * 72)
    print(f"{'CASE':<12}{'RESULT':<12}{'EXPECTED':<12}{'':<3}SCENARIO")
    print("-" * 72)
    for case, (want, scenario) in EXPECTED.items():
        got = results.get(case, "MISSING")
        ok = got == want
        if not ok:
            failures.append(f"{case}: got {got}, expected {want}")
        print(f"{case:<12}{got:<12}{want:<12}{'OK ' if ok else 'FAIL':<3}{scenario}")

    print("-" * 72)
    checks = []

    # the filtered-out row must never be touched
    checks.append(("filtered-out row 123021 never processed",
                   "SOC123021" not in results))

    # no reply may be drafted for the case whose search misfired
    checks.append(("no reply drafted for the failed search (610888)",
                   all(num != "610888" for num, _ in app.replies_started)))

    # the duplicate case must reply to the NEWEST mail
    newest = [subj for num, subj in app.replies_started if num == "610999"]
    checks.append(("replied to the newest of 2 mails (610999)",
                   newest == ["SOC610999 latest update"]))

    # the discarded draft must not be counted as sent
    checks.append(("discarded draft not counted as sent (610529)",
                   all(m.num != "610529" for m in app.sent_items)))

    # every genuine send landed in Sent Items
    checks.append(("sent replies verified in Sent Items",
                   {m.num for m in app.sent_items} == {"610321", "610454", "610999"}))

    # the '...' fallback really was exercised
    checks.append(("Reply all found via the '...' menu (610454)",
                   any(num == "610454" for num, _ in app.replies_started)))

    # screenshot evidence for not-found cases
    checks.append(("screenshots saved for not-found cases",
                   len([s for s in app.shots if "notfound" in s]) >= 2))

    # output files
    out_csv = summary["out_csv"]
    csv_rows = list(csv.DictReader(open(out_csv, encoding="utf-8-sig")))
    checks.append((f"results CSV written ({len(csv_rows)} rows)",
                   len(csv_rows) == len(cases)))

    xlsx = summary["xlsx"]
    red_rows, green_rows, sheet_cases = [], [], []
    if xlsx and Path(xlsx).exists():
        from openpyxl import load_workbook
        ws = load_workbook(xlsx).active
        for row in ws.iter_rows(min_row=2):
            rgb = (row[0].fill.start_color.rgb or "")[-6:] if row[0].fill else ""
            case = str(row[0].value)
            sheet_cases.append(case)
            if rgb == "FFC7CE":
                red_rows.append(case)
            elif rgb == "C6EFCE":
                green_rows.append(case)
    checks.append(("cases needing follow-up shown red in the review sheet",
                   sorted(red_rows) == ["610777", "610888"], str(red_rows)))
    checks.append(("replied cases shown green",
                   sorted(green_rows) == ["610321", "610454", "610999"],
                   str(green_rows)))
    # rows the analyst filtered out in Excel stay out of the review sheet
    checks.append(("review sheet covers exactly the filtered-to cases",
                   sorted(sheet_cases) == sorted(c for c in cases),
                   str(sheet_cases)))

    # post-send pause: after each send except the final case (610999 is last)
    pauses = len([l for l in ui.lines if "before the next case" in l])
    checks.append((f"post-send pause applied ({pauses}x: after each send "
                   "except the last case)", pauses == 2))

    # tracker statuses reached the UI
    checks.append(("live status updates emitted to the GUI tracker",
                   any(s == "WORKING" for _, s in ui.updates)
                   and any(s == "REVIEW" for _, s in ui.updates)))

    for item in checks:
        label, ok = item[0], item[1]
        note = f"  ({item[2]})" if len(item) > 2 and not ok else ""
        print(f"[{'OK ' if ok else 'FAIL'}] {label}{note}")
        if not ok:
            failures.append(label + note)

    print("-" * 72)
    print(f"Replies drafted : {len(app.replies_started)}")
    print(f"Actually sent   : {len(app.sent_items)}")
    print(f"Run time        : {elapsed:.1f}s (simulated delays)")
    print(f"Outputs         : {Path(out_csv).name}"
          + (f", {Path(xlsx).name}" if xlsx else ""))
    print("=" * 72)
    if failures:
        print(f"AUDIT FAILED - {len(failures)} problem(s):")
        for f in failures:
            print("   -", f)
    else:
        print("AUDIT PASSED - every scenario behaved as specified.")
    print("=" * 72)

    shutil.rmtree(workdir, ignore_errors=True)
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
