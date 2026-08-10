#!/usr/bin/env python3
"""Data-integrity tests.

Every case here is a way the tool could quietly mislead an analyst - a case
dropped without being counted, a row coloured as though it were handled, or
a whole run's results lost. Each was found by review.

    python test_data_integrity.py
"""

import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parent))
import soc_followup as core
from openpyxl import Workbook, load_workbook

PASS, FAIL = [], []
TMP = Path(tempfile.mkdtemp(prefix="soc_integrity_"))


def check(label, ok, detail=""):
    (PASS if ok else FAIL).append(label)
    print(f"[{'OK ' if ok else 'FAIL'}] {label}")
    if detail and not ok:
        print(f"        {detail}")


def sheet(rows, headers=("id", "name"), name="s.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    path = TMP / name
    wb.save(path)
    return path


print("=" * 72)
print("DATA INTEGRITY")
print("=" * 72)

# --- the right column must win -------------------------------------------
print("\n-- choosing the case column --")
p = sheet([["6074563", "alert", ""], ["6074420", "alert", ""],
           ["6074294", "alert", ""], ["6074424", "alert", "999999"]],
          headers=("id", "name", "notes"))
col, cases = core.extract_cases(p)
check("a mostly-empty column does not beat the real id column",
      col == "id" and len(cases) == 4, f"chose {col!r} -> {cases}")

# --- repeated headings ---------------------------------------------------
p = sheet([["6074563", "alert", "junk"]], headers=("id", "name", "id"))
col, cases = core.extract_cases(p)
check("a repeated column heading does not shadow the first one",
      cases == ["6074563"], f"chose {col!r} -> {cases}")

# --- nothing may be dropped silently -------------------------------------
print("\n-- nothing dropped without being counted --")
p = sheet([["6074563", "ok"], ["", "blank id"], ["NOPE", "unreadable"],
           ["6074563", "duplicate"], ["6074420", "ok"]])
rep = core.preflight(p)
check("a blank case cell is counted, not silently dropped",
      len(rep["blanks"]) == 1, f"blanks={rep['blanks']}")
check("an unreadable case cell is reported with its row number",
      rep["unreadable"] == [(4, "NOPE")], str(rep["unreadable"]))
check("a duplicate is reported with its row number",
      rep["duplicates"] == [(5, "6074563")], str(rep["duplicates"]))
check("every row is accounted for",
      len(rep["cases"]) + len(rep["blanks"]) + len(rep["unreadable"])
      + len(rep["duplicates"]) == rep["rows_read"],
      f"{len(rep['cases'])} + blanks/unreadable/dupes vs {rep['rows_read']}")

# --- the review sheet must not overstate what happened -------------------
print("\n-- review sheet honesty --")
p = sheet([["700001", "handled"], ["700001", "same case again"],
           ["700002", "handled"]])
stats = {}
col, _ = core.extract_cases(p, stats=stats)
out = TMP / "review.xlsx"
core.write_status_xlsx(p, col, {"700001": "SENT", "700002": "SENT"}, out,
                       stats.get("duplicate_lines"))
ws = load_workbook(out).active


def status_col(sheet):
    """Locate the status column by name - a 'Why' column follows it."""
    for cell in sheet[1]:
        if str(cell.value).strip().lower() == "follow-up status":
            return cell.column - 1        # 0-based index into the row tuple
    return -1


sc = status_col(ws)
rows = [(str(r[0].value), r[sc].value,
         ((r[0].fill.start_color.rgb or "")[-6:] if r[0].fill else ""))
        for r in ws.iter_rows(min_row=2)]
dup_row = rows[1]
check("a duplicate row is not painted as though it were replied to",
      dup_row[1] == "Duplicate row" and dup_row[2] != "C6EFCE", str(dup_row))
check("the genuine rows are still green",
      rows[0][2] == "C6EFCE" and rows[2][2] == "C6EFCE", str(rows))

out2 = TMP / "review2.xlsx"
core.write_status_xlsx(p, col, {"700001": "SENT_UNVERIFIED"}, out2)
ws2 = load_workbook(out2).active
first = next(ws2.iter_rows(min_row=2))
sc2 = status_col(ws2)
check("a reply that could not be confirmed is not shown as plain green",
      "unconfirmed" in str(first[sc2].value).lower()
      and (first[0].fill.start_color.rgb or "")[-6:] != "C6EFCE",
      str(first[sc2].value))

# --- the sheet must explain itself ---------------------------------------
print("\n-- the sheet says why --")
out3 = TMP / "review3.xlsx"
core.write_status_xlsx(
    p, col,
    {"700001": "NOT_FOUND", "700002": "SKIPPED"}, out3,
    stats.get("duplicate_lines"),
    {"700001": "Searched 5 ways - no mail mentions 700001",
     "700002": "Draft was opened; the analyst skipped it without sending"})
ws3 = load_workbook(out3).active
headers3 = [str(c.value) for c in ws3[1]]
body3 = list(ws3.iter_rows(min_row=2))
check("there is a 'Why' column after the status",
      headers3[-2:] == ["Follow-up status", "Why"], str(headers3[-2:]))
check("a not-found row explains how it searched",
      "Searched 5 ways" in str(body3[0][-1].value), str(body3[0][-1].value))
check("a skipped row explains that the analyst skipped it",
      "skipped it without sending" in str(body3[2][-1].value),
      str(body3[2][-1].value))
check("a duplicate row explains itself too",
      "earlier in the sheet" in str(body3[1][-1].value),
      str(body3[1][-1].value))

# --- results survive a mid-run failure -----------------------------------
print("\n-- results survive a crash --")


class UI:
    def log(self, m): pass
    def case_update(self, *a): pass
    def ask(self, *a, **k): return ""
    def ask_or_watch(self, *a, **k): return ""


def boom(cases, opts, ui, results=None):
    results.append({"case": "SOC700001", "status": "SENT", "detail": "done"})
    results.append({"case": "SOC700002", "status": "NOT_FOUND", "detail": "-"})
    raise RuntimeError("browser died")


core.SCRIPT_DIR = TMP
original = core._run_web
core._run_web = boom
try:
    summary = core.run_followups(sheet([["700001", "a"], ["700002", "b"]]),
                                 "id", ["700001", "700002"],
                                 SimpleNamespace(no_pause=True), UI())
    kept = len(summary["results"])
    written = Path(summary["out_csv"]).exists()
finally:
    core._run_web = original

check("cases finished before a crash are kept", kept == 2, f"kept {kept}")
check("the results file is still written after a crash", written)

# --- CSV from Excel ------------------------------------------------------
print("\n-- encodings --")
csv_path = TMP / "cp1252.csv"
csv_path.write_bytes("id,name\n6074563,Caf\xe9 alert\n".encode("cp1252"))
col, cases = core.extract_cases(csv_path)
check("an Excel-saved CSV (cp1252) is readable", cases == ["6074563"],
      str(cases))

print("\n" + "-" * 72)
print(f"{len(PASS)} passed, {len(FAIL)} failed")
for f in FAIL:
    print("   FAILED:", f)
print("=" * 72)
sys.exit(1 if FAIL else 0)
